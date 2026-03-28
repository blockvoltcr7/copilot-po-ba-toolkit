"""
AI/ML Financial Services Regulatory & Compliance Landscape 2025-2026
Generates a multi-sheet Excel workbook with detailed regulatory data.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
import os

wb = Workbook()

# Color palette
NAVY = "1B2A4A"
DARK_BLUE = "065A82"
TEAL = "1C7293"
WHITE = "FFFFFF"
LIGHT_BG = "F0F4F8"
LIGHT_BLUE = "DBEAFE"
CRITICAL_RED = "DC2626"
HIGH_ORANGE = "EA580C"
MEDIUM_YELLOW = "CA8A04"
LOW_GREEN = "16A34A"
LIGHT_RED = "FEE2E2"
LIGHT_ORANGE = "FFF7ED"
LIGHT_YELLOW = "FEF9C3"
LIGHT_GREEN = "DCFCE7"

header_font = Font(name="Arial", bold=True, color=WHITE, size=10)
header_fill = PatternFill("solid", fgColor=NAVY)
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
body_font = Font(name="Arial", size=9)
body_align = Alignment(vertical="top", wrap_text=True)
center_align = Alignment(horizontal="center", vertical="top", wrap_text=True)
thin_border = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1"),
)
alt_fill = PatternFill("solid", fgColor=LIGHT_BG)

def style_header(ws, row, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

def style_row(ws, row, col_count, alt=False, center_cols=None):
    center_cols = center_cols or []
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = body_font
        cell.alignment = center_align if col in center_cols else body_align
        cell.border = thin_border
        if alt:
            cell.fill = alt_fill

def add_title(ws, title, subtitle=None):
    ws.merge_cells("A1:N1")
    c = ws.cell(row=1, column=1, value=title)
    c.font = Font(name="Georgia", bold=True, size=16, color=NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 35
    if subtitle:
        ws.merge_cells("A2:N2")
        c2 = ws.cell(row=2, column=1, value=subtitle)
        c2.font = Font(name="Arial", size=10, color="64748B", italic=True)
        c2.alignment = Alignment(horizontal="left")
        ws.row_dimensions[2].height = 22

# ============================================================
# SHEET 1 — Master Regulatory Register
# ============================================================
ws1 = wb.active
ws1.title = "Regulatory Register"
add_title(ws1, "AI/ML Financial Services — Regulatory & Compliance Register",
          "Comprehensive landscape as of 2025-2026 | For wealth management compliance planning")

headers = [
    "ID", "Regulation / Guidance", "Full Citation", "Issuing Body",
    "Effective Date / Status", "Risk Level", "Impact Area",
    "Description", "Key Requirements",
    "Compliance Actions Required", "Deadline",
    "Wealth Mgmt Relevance", "Enforcement Examples", "Source URL"
]
widths = [5, 28, 35, 14, 18, 12, 20, 45, 50, 50, 16, 40, 40, 45]

start_row = 4
for i, (h, w) in enumerate(zip(headers, widths), 1):
    ws1.cell(row=start_row, column=i, value=h)
    ws1.column_dimensions[get_column_letter(i)].width = w
style_header(ws1, start_row, len(headers))

regulations = [
    [
        "REG-01",
        "SEC Proposed PDA/AI Conflicts Rule",
        "Proposed Rule: Conflicts of Interest Associated with the Use of Predictive Data Analytics, Release No. IA-6399 / 34-97990 (July 2023)",
        "SEC",
        "Proposed (Jul 2023); Comment period closed Oct 2023; Final rule pending",
        "Critical",
        "Disclosure; Conflicts of Interest",
        "Requires broker-dealers and investment advisers to evaluate, identify, and eliminate or neutralize conflicts of interest associated with the use of predictive data analytics, AI, ML, and similar covered technologies in investor interactions. Disclosure alone is insufficient.",
        "1) Identify conflicts from AI/PDA use in investor interactions\n2) Eliminate or neutralize (not merely disclose) identified conflicts\n3) Develop written policies and procedures\n4) Conduct annual reviews and testing\n5) Maintain books and records of evaluations",
        "1) Inventory all AI/ML tools used in client interactions\n2) Map conflicts of interest for each tool\n3) Implement conflict elimination/neutralization controls\n4) Draft AI-specific compliance policies\n5) Establish annual review cadence\n6) Build recordkeeping infrastructure",
        "18 months post-adoption (estimated 2026)",
        "Directly impacts robo-advisors, AI-driven portfolio recommendations, client engagement chatbots, and any AI tool influencing investment decisions for wealth management clients.",
        "N/A — proposed rule. SEC initiated AI sweep examinations in late 2023 requesting information from multiple advisers.",
        "https://www.sec.gov/newsroom/press-releases/2023-140"
    ],
    [
        "REG-02",
        "SEC Marketing Rule (AI-Washing Enforcement)",
        "Investment Advisers Act Rule 206(4)-1 (Marketing Rule); Effective Nov 4, 2022; Enforcement actions under Sections 206(2), 206(4), Rules 206(4)-1 and 206(4)-7",
        "SEC",
        "Effective (Nov 2022); Active enforcement since Mar 2024",
        "Critical",
        "Disclosure; Consumer Protection",
        "Prohibits misleading statements in adviser marketing, including unsubstantiated AI capability claims ('AI-washing'). All marketing statements about AI must be substantiated by evidence reflecting actual practices.",
        "1) All AI claims in marketing must be substantiated\n2) No untrue statements of material fact in advertisements\n3) Written compliance policies for AI marketing claims\n4) Principles-based prohibitions on exaggeration/omission\n5) Fair presentation of AI capabilities",
        "1) Audit all marketing materials for AI capability claims\n2) Document substantiation evidence for each AI claim\n3) Update compliance policies for AI-specific marketing review\n4) Implement pre-publication review process for AI content\n5) Train marketing staff on AI-washing risks",
        "Immediate — already in effect",
        "Critical for wealth managers marketing AI-driven investment strategies, robo-advisory capabilities, or AI-enhanced portfolio management. Claims must match actual deployed technology.",
        "Delphia (USA) Inc. & Global Predictions Inc. (Mar 2024): $400K combined penalties for false AI claims. Rimar Capital USA (Oct 2024): enforcement for misrepresenting AI trading capabilities.",
        "https://www.sec.gov/newsroom/press-releases/2024-36"
    ],
    [
        "REG-03",
        "SEC Regulation Best Interest (Reg BI) — AI Implications",
        "17 CFR § 240.15l-1 (Regulation Best Interest); Effective June 30, 2020",
        "SEC",
        "Effective (Jun 2020); AI-specific focus in 2024-2025 exam priorities",
        "Critical",
        "Conflicts of Interest; Consumer Protection",
        "Reg BI requires broker-dealers to act in retail clients' best interest when making recommendations. AI-generated recommendations are fully subject to Reg BI's disclosure, care, conflict of interest, and compliance obligations.",
        "1) AI recommendations must meet best interest standard\n2) Disclose role and limitations of AI tools to clients\n3) Ensure AI recommendations are appropriate per client profile\n4) Identify/mitigate conflicts from AI favoring proprietary products\n5) Maintain audit trails for AI-driven decisions",
        "1) Review AI recommendation engines for Reg BI compliance\n2) Update Form CRS disclosures to address AI use\n3) Implement conflict-of-interest testing for AI models\n4) Build explainability into AI recommendation logic\n5) Train compliance staff on AI-specific Reg BI obligations",
        "Immediate — ongoing obligation",
        "Directly governs AI-powered recommendations, robo-advisory, and automated portfolio allocation in wealth management. SEC examiners explicitly prioritizing this area in 2025.",
        "SEC 2025 examination priorities explicitly list AI-driven recommendations and Reg BI compliance as focus areas.",
        "https://www.sec.gov/about/divisions-offices/division-trading-markets/regulation-best-interest-form-crs-related-interpretations"
    ],
    [
        "REG-04",
        "FINRA Regulatory Notice 24-09 (AI in Securities Industry)",
        "FINRA Regulatory Notice 24-09: Artificial Intelligence (AI) (June 2024)",
        "FINRA",
        "Effective (Jun 2024); Active guidance",
        "High",
        "Model Risk; Disclosure; Consumer Protection",
        "Reminds member firms that existing FINRA rules — including supervision (Rule 3110), communications (Rule 2210), and recordkeeping — apply technology-neutrally to AI tools. Firms must integrate AI into existing supervisory systems with proper governance.",
        "1) Apply Rule 3110 supervisory requirements to AI tools\n2) Rule 2210 compliance for AI-generated communications\n3) Full recordkeeping of AI-generated content and decisions\n4) Technology governance: model risk, data privacy, accuracy\n5) Vendor due diligence for third-party AI systems",
        "1) Update supervisory procedures to cover AI tools\n2) Establish AI content approval workflows per Rule 2210\n3) Implement AI output recordkeeping and archival\n4) Conduct AI vendor risk assessments\n5) Train supervisors on AI oversight responsibilities\n6) Create AI use case inventory",
        "Immediate — existing rules apply now",
        "Applies to all broker-dealer activities using AI: chatbots for client service, AI-generated research reports, AI-assisted trade surveillance, and automated communications.",
        "FINRA examinations are actively reviewing AI governance and supervisory systems in 2024-2025 annual reviews.",
        "https://www.finra.org/rules-guidance/notices/24-09"
    ],
    [
        "REG-05",
        "FINRA 2025 Annual Regulatory Oversight Report (AI Focus)",
        "FINRA 2025 Annual Regulatory Oversight Report — AI/Emerging Technology Section (Feb 2025)",
        "FINRA",
        "Published Feb 2025; Active guidance",
        "High",
        "Model Risk; Vendor Risk",
        "Expands AI supervisory expectations: real-time monitoring of AI outputs, enhanced documentation standards, specialized supervisor training, and enterprise-level AI governance. Treats AI systems as supervised entities comparable to human staff.",
        "1) Real-time monitoring of AI outputs\n2) Enhanced documentation and audit trails\n3) Specialized AI training for supervisors\n4) Enterprise-level AI governance framework\n5) AI model inventory management\n6) Risk assessment for evolving AI use cases",
        "1) Implement real-time AI output monitoring capabilities\n2) Establish AI centers of excellence\n3) Develop AI-specific supervisor training programs\n4) Build comprehensive AI model/use-case inventories\n5) Define prohibited AI activities and guidelines\n6) Create ongoing AI risk assessment processes",
        "Align by end of 2025",
        "Wealth management firms must demonstrate robust AI governance comparable to oversight of human employees, especially for client-facing AI applications.",
        "2025 Oversight Report highlights ongoing examination focus on AI governance maturity across member firms.",
        "https://www.finra.org/rules-guidance/guidance/reports/2025-finras-annual-regulatory-oversight-report"
    ],
    [
        "REG-06",
        "Federal Reserve SR 11-7 (Model Risk Management)",
        "SR Letter 11-7: Supervisory Guidance on Model Risk Management (April 4, 2011); Companion: OCC Bulletin 2011-12",
        "Federal Reserve / OCC",
        "Effective (Apr 2011); Fully applies to AI/ML — heightened 2024-2025 examiner expectations",
        "Critical",
        "Model Risk",
        "Foundational MRM framework requiring model inventory, independent validation, ongoing monitoring, and board-level governance. Applies with full force to AI/ML models including neural networks, LLMs, and GenAI. Examiners now expect AI-specific controls.",
        "1) Comprehensive model inventory (including shadow AI and vendor models)\n2) Three lines of defense: developers, validators, internal audit\n3) Independent validation: conceptual soundness, ongoing monitoring, outcomes analysis\n4) AI-specific: explainability, adversarial robustness, reproducibility\n5) Board/senior management oversight of aggregate model risk\n6) Data drift and concept drift monitoring",
        "1) Expand model inventory to capture all AI/ML models\n2) Develop AI-specific validation procedures\n3) Implement ongoing monitoring for data/model drift\n4) Document hyperparameters, training data, and assumptions\n5) Establish bias testing and fairness analysis protocols\n6) Build vendor/third-party model oversight process\n7) Ensure board reporting on AI model risk",
        "Immediate — ongoing obligation; AI-specific enhancements expected by mid-2025",
        "Core framework for all AI/ML models in wealth management: portfolio optimization, risk scoring, client segmentation, fraud detection, and AML models.",
        "Examiners frequently flag: incomplete model inventories (missing shadow/vendor AI), insufficient explainability documentation, weak ongoing monitoring for drift or bias.",
        "https://www.federalreserve.gov/supervisionreg/srletters/sr1107.htm"
    ],
    [
        "REG-07",
        "OCC Bulletin 2023-17 (Third-Party Risk Management)",
        "OCC Bulletin 2023-17: Interagency Guidance on Third-Party Relationships: Risk Management (June 2023); Rescinds OCC 2013-29 and 2020-10",
        "OCC / Fed / FDIC",
        "Effective (Jun 2023)",
        "High",
        "Vendor Risk",
        "Standardized interagency guidance for third-party risk management covering the full lifecycle: planning, due diligence, selection, contract negotiation, ongoing monitoring, and termination. Banks retain ultimate responsibility even when outsourcing to AI vendors.",
        "1) Life-cycle approach to third-party management\n2) Proportional oversight based on risk/criticality\n3) Due diligence before and during vendor relationships\n4) Contract requirements for AI service providers\n5) Ongoing monitoring of vendor performance and risks\n6) Termination planning and data return provisions",
        "1) Map all AI vendor relationships and criticality tiers\n2) Conduct comprehensive due diligence on AI vendors\n3) Review and update contracts with AI-specific provisions\n4) Establish ongoing vendor performance monitoring\n5) Develop concentration risk assessments for AI providers\n6) Create contingency/exit plans for critical AI vendors\n7) Ensure subcontractor oversight requirements",
        "Immediate — already in effect",
        "Essential for wealth managers using third-party AI platforms, robo-advisory providers, AI analytics vendors, and cloud-based AI services.",
        "OCC Bulletin 2024-11 provides additional community bank guidance. Examiners scrutinize both outsourcing process and technical controls for AI models.",
        "https://www.occ.gov/news-issuances/bulletins/2023/bulletin-2023-17.html"
    ],
    [
        "REG-08",
        "CFPB Circular 2022-03 (Adverse Action Notices & AI)",
        "Consumer Financial Protection Circular 2022-03: Adverse Action Notification Requirements in Connection with Credit Decisions Based on Complex Algorithms (May 2022)",
        "CFPB",
        "Effective (May 2022)",
        "Critical",
        "Fair Lending; Consumer Protection",
        "Mandates that creditors must provide specific, accurate reasons for adverse credit actions under ECOA/Regulation B, regardless of AI/algorithm complexity. Black-box models that cannot generate specific reasons are unlawful for credit decisioning.",
        "1) Provide specific principal reasons for adverse actions — not generic statements\n2) Reasons must accurately reflect actual factors from the AI model\n3) Complexity of AI is not an excuse for inadequate explanations\n4) Post-hoc explainability tools allowed but must be validated\n5) If model cannot provide specific reasons, it cannot be used",
        "1) Evaluate all AI credit models for adverse action explainability\n2) Implement validated explainability techniques (SHAP, LIME, etc.)\n3) Test adverse action reason codes for accuracy and specificity\n4) Update adverse action notice templates\n5) Document model explainability methodology\n6) Train compliance staff on AI explanation requirements",
        "Immediate — already in effect",
        "Directly impacts any wealth management firm offering credit products (margin lending, securities-based lending, credit lines) using AI-driven underwriting.",
        "CFPB examination findings in 2024-2025 cite failures in compliance management systems and insufficient testing for less discriminatory alternatives in AI credit models.",
        "https://www.consumerfinance.gov/compliance/circulars/circular-2022-03-adverse-action-notification-requirements-in-connection-with-credit-decisions-based-on-complex-algorithms/"
    ],
    [
        "REG-09",
        "ECOA / Regulation B — Fair Lending & AI Bias",
        "Equal Credit Opportunity Act, 15 U.S.C. § 1691 et seq.; Regulation B, 12 CFR Part 1002",
        "CFPB / DOJ",
        "Effective (1974; ongoing); AI-specific enforcement focus 2024-2025",
        "Critical",
        "Fair Lending",
        "Prohibits credit discrimination based on protected characteristics. AI/ML credit models must be tested for disparate impact and disparate treatment. CFPB 2025 proposed amendments to Regulation B would eliminate disparate impact liability under ECOA, but disparate treatment remains actionable. State laws and FHA still apply.",
        "1) No discrimination based on protected characteristics in AI models\n2) Test AI models for both disparate impact and disparate treatment\n3) Explore less discriminatory alternatives that maintain accuracy\n4) Document business necessity for model features\n5) Regular fair lending analyses of AI model outcomes\n6) Adequate compliance management systems",
        "1) Conduct disparate impact testing on all AI credit models\n2) Implement bias detection and mitigation frameworks\n3) Test for less discriminatory alternatives\n4) Document business justifications for model variables\n5) Establish regular fair lending review cycles\n6) Monitor CFPB Reg B proposed amendments\n7) Ensure state-level fair lending compliance",
        "Immediate — ongoing obligation",
        "Critical for any wealth management firm using AI in credit decisions: margin lending, SBL, personal loans. Even with CFPB's proposed narrowing of disparate impact, firms should maintain robust testing programs due to state law exposure.",
        "CFPB 2024 Fair Lending Report: advanced credit scoring models led to disproportionately negative outcomes for Black and Hispanic applicants. 2025 proposed Reg B amendments to eliminate disparate impact under ECOA.",
        "https://www.consumerfinance.gov/data-research/research-reports/fair-lending-report-of-the-consumer-financial-protection-bureau-2024/"
    ],
    [
        "REG-10",
        "GLBA — Data Privacy for AI in Financial Services",
        "Gramm-Leach-Bliley Act (GLBA), 15 U.S.C. §§ 6801-6809; FTC Safeguards Rule, 16 CFR Part 314 (updated effective Jun 2023)",
        "FTC / Federal Banking Agencies",
        "Effective (1999; FTC Safeguards Rule updated Jun 2023)",
        "High",
        "Data Privacy",
        "Requires financial institutions to safeguard customer NPI (nonpublic personal information), provide privacy notices, and implement information security programs. AI systems processing customer data must comply with GLBA data handling, sharing, and security requirements.",
        "1) Privacy notices explaining AI data usage\n2) Safeguards for NPI used in AI training/inference\n3) Opt-out rights for data sharing with AI vendors\n4) Information security program covering AI systems\n5) Access controls for AI model training data\n6) FTC Safeguards Rule: written security plan, designated CISO, risk assessments",
        "1) Audit AI data pipelines for NPI handling compliance\n2) Update privacy notices to disclose AI data use\n3) Implement data minimization for AI training data\n4) Review vendor contracts for AI data sharing provisions\n5) Ensure GLBA security requirements cover AI infrastructure\n6) Conduct risk assessments of AI data flows\n7) Maintain written information security plan for AI systems",
        "Immediate — ongoing obligation; FTC Safeguards Rule compliance required",
        "Governs how wealth managers can use client financial data for AI model training, client profiling, and personalization. Restricts sharing client NPI with AI technology vendors.",
        "FTC enforcement actions on Safeguards Rule compliance. Some states (Montana, Connecticut) reducing GLBA exemptions in 2025, creating dual compliance requirements.",
        "https://www.ftc.gov/business-guidance/privacy-security/gramm-leach-bliley-act"
    ],
    [
        "REG-11",
        "CCPA/CPRA — California AI Data Privacy",
        "California Consumer Privacy Act (CCPA) / California Privacy Rights Act (CPRA), Cal. Civ. Code §§ 1798.100-1798.199.100; Effective Jan 1, 2023 (CPRA amendments)",
        "California AG / CPPA",
        "Effective (Jan 2020 CCPA; Jan 2023 CPRA amendments)",
        "High",
        "Data Privacy",
        "Grants California consumers rights to access, delete, correct, and opt out of sale/sharing of personal information. CPRA adds automated decision-making provisions and data minimization requirements. Financial institutions have data-level (not entity-level) GLBA exemptions.",
        "1) Consumer rights: access, deletion, correction, opt-out\n2) Automated decision-making transparency requirements\n3) Data minimization for AI processing\n4) Right to opt out of automated profiling\n5) Data-level GLBA exemption only (not entity-level)\n6) Risk assessments for high-risk processing including AI",
        "1) Map all personal data used in AI systems\n2) Implement consumer rights fulfillment for AI-processed data\n3) Provide automated decision-making disclosures\n4) Enable opt-out mechanisms for AI profiling\n5) Conduct data protection impact assessments for AI\n6) Review GLBA exemption applicability per data type\n7) Monitor CPPA rulemaking on automated decision-making",
        "Immediate — already in effect; CPPA automated decision-making rules pending",
        "Affects wealth managers with California clients using AI for client profiling, investment recommendations, risk scoring, and marketing personalization.",
        "CPPA actively developing regulations for automated decision-making technology (ADMT). Financial institutions face dual GLBA/CCPA compliance requirements in California.",
        "https://oag.ca.gov/privacy/ccpa"
    ],
    [
        "REG-12",
        "NIST AI Risk Management Framework (AI RMF 1.0)",
        "NIST AI 100-1: Artificial Intelligence Risk Management Framework 1.0 (January 2023); GenAI Profile: NIST AI 600-1 (July 2024)",
        "NIST",
        "Published Jan 2023; GenAI Profile Jul 2024; Voluntary framework",
        "Medium",
        "Model Risk; Vendor Risk",
        "Voluntary, sector-agnostic framework for managing AI risks across Govern, Map, Measure, and Manage functions. Financial Services AI RMF (FS AI RMF) adapts with 230+ control objectives for financial institutions. Becoming de facto standard for US AI governance.",
        "1) Govern: Establish AI governance structures and accountability\n2) Map: Identify and classify AI risks in context\n3) Measure: Assess and analyze AI risks quantitatively\n4) Manage: Treat, monitor, and communicate AI risks\n5) GenAI Profile: Address generative AI-specific risks\n6) FS AI RMF: Financial sector-specific controls",
        "1) Adopt NIST AI RMF as baseline governance framework\n2) Implement Govern/Map/Measure/Manage functions\n3) Align with FS AI RMF sector-specific controls\n4) Integrate with existing MRM (SR 11-7) programs\n5) Apply GenAI Profile for LLM/GenAI deployments\n6) Benchmark current AI maturity and set targets\n7) Use framework for vendor evaluation criteria",
        "Voluntary; recommended adoption by Q2 2025 for regulatory readiness",
        "Increasingly referenced by regulators and examiners as expected standard. Wealth managers adopting proactively gain regulatory credibility and structured approach to AI risk governance.",
        "Referenced by multiple federal agencies as expected best practice. Financial Services AI RMF developed by Cyber Risk Institute with industry input.",
        "https://www.nist.gov/itl/ai-risk-management-framework"
    ],
    [
        "REG-13",
        "EU AI Act — Extraterritorial Impact on US Firms",
        "Regulation (EU) 2024/1689: EU Artificial Intelligence Act (Adopted Jun 2024; Full enforcement Aug 2026)",
        "European Parliament / Council",
        "Adopted Jun 2024; Phased: Feb 2025 (prohibited), Aug 2025 (GPAI), Aug 2026 (high-risk full enforcement)",
        "High",
        "Model Risk; Disclosure; Fair Lending; Data Privacy",
        "Classifies AI systems by risk level. Financial services applications (credit scoring, fraud detection, AML, algorithmic trading) are high-risk. Applies extraterritorially to US firms serving EU clients or placing AI products in EU markets. Penalties up to 7% global annual turnover.",
        "1) High-risk AI: conformity assessments, quality management systems\n2) Detailed technical documentation and logging\n3) Human oversight requirements\n4) Transparency and explainability for users\n5) High-quality unbiased training data\n6) Regular monitoring and incident reporting\n7) AI system inventory and registration",
        "1) Map all AI systems serving EU clients\n2) Classify systems per EU AI Act risk tiers\n3) Conduct conformity assessments for high-risk AI\n4) Implement quality management systems\n5) Ensure documentation meets EU requirements\n6) Build human oversight into AI workflows\n7) Prepare for EU AI Office registration\n8) Align with DORA operational resilience requirements",
        "Aug 2026 (high-risk full enforcement); Begin preparation immediately",
        "US wealth managers with EU clients, branches, or EU market presence must comply. Credit scoring, KYC, fraud detection, and algorithmic trading systems are high-risk. Brussels effect may drive global standards adoption.",
        "First prohibited AI practices enforceable Feb 2025. GPAI (foundation model) rules from Aug 2025. Full high-risk system rules from Aug 2026.",
        "https://eur-lex.europa.eu/eli/reg/2024/1689/oj"
    ],
    [
        "REG-14",
        "OCC Bulletin 2011-12 (Model Risk Management)",
        "OCC Bulletin 2011-12: Supervisory Guidance on Model Risk Management (April 2011); Companion to SR 11-7",
        "OCC",
        "Effective (Apr 2011); Heightened AI/ML focus 2024-2025",
        "Critical",
        "Model Risk",
        "OCC's companion to SR 11-7 establishing model risk management expectations for national banks. Defines 'model' broadly enough to encompass AI/ML. Requires sound development, validation, and governance practices. Fair lending implications for AI credit models explicitly within scope.",
        "1) Sound model development practices with documentation\n2) Effective model validation — independent from development\n3) Model governance including roles, responsibilities, policies\n4) Model use controls and change management\n5) Board and senior management oversight\n6) Fair lending testing for credit models",
        "1) Ensure AI/ML models are captured under MRM governance\n2) Document development methodology for AI models\n3) Establish independent AI model validation\n4) Implement change management for AI model updates/retraining\n5) Include fair lending analysis in AI model validation\n6) Report AI model risk to board/risk committee",
        "Immediate — ongoing obligation",
        "Governs all AI/ML models at national banks used in wealth management: credit risk, portfolio optimization, client risk profiling, AML/KYC.",
        "OCC examinations focus on model inventory completeness, validation independence, and fair lending testing for AI/ML models.",
        "https://www.occ.gov/news-issuances/bulletins/2011/bulletin-2011-12.html"
    ],
    [
        "REG-15",
        "SEC 2025 Examination Priorities (AI Focus)",
        "SEC Division of Examinations: 2025 Examination Priorities (Published Oct 2024)",
        "SEC",
        "Active for 2025 examination cycle",
        "High",
        "Model Risk; Disclosure; Conflicts of Interest",
        "SEC's 2025 exam priorities explicitly identify AI/emerging technology as a focus area. Examiners will assess how registrants are meeting existing obligations when adopting AI for investment advice, trading, compliance, and marketing.",
        "1) AI governance and oversight frameworks\n2) Disclosures about AI use to clients\n3) Compliance with Marketing Rule for AI claims\n4) Reg BI/fiduciary compliance in AI recommendations\n5) Cybersecurity and data privacy for AI systems\n6) Recordkeeping of AI-assisted decisions",
        "1) Prepare for SEC AI-focused examinations\n2) Document AI governance structure and policies\n3) Ensure all AI disclosures are current and accurate\n4) Conduct self-assessment against exam priorities\n5) Prepare AI-related documentation packages\n6) Brief senior management on AI examination risks",
        "Ongoing through 2025 examination cycle",
        "Wealth management firms should expect targeted AI-related examination inquiries. Preparation is essential for firms using AI in any client-facing or investment decision capacity.",
        "SEC initiated AI sweep exams in late 2023. 2025 priorities signal continued escalation of AI-focused regulatory scrutiny.",
        "https://www.sec.gov/about/divisions-offices/division-examinations/examination-priorities"
    ],
    [
        "REG-16",
        "State AI Legislation (Multi-State Compliance)",
        "Colorado SB 24-205: AI Consumer Protections Act (Signed May 2024; Effective Feb 1, 2026); Illinois HB 3773: AI Video Interview Act; Various state proposals",
        "State Legislatures",
        "Colorado effective Feb 2026; Various other states pending",
        "Medium",
        "Fair Lending; Consumer Protection; Disclosure",
        "Growing wave of state-level AI legislation. Colorado's AI Act requires deployers of high-risk AI to conduct impact assessments, provide disclosures, and implement governance programs. Other states pursuing similar legislation creating patchwork compliance landscape.",
        "1) Impact assessments for high-risk AI systems\n2) Consumer notifications when interacting with AI\n3) Opt-out rights for algorithmic profiling\n4) Bias testing and documentation\n5) Governance programs for AI deployers\n6) Annual compliance reporting",
        "1) Monitor state AI legislation in all operating jurisdictions\n2) Conduct impact assessments per Colorado AI Act\n3) Implement consumer notification mechanisms\n4) Develop state-by-state compliance matrices\n5) Prepare for patchwork regulatory landscape\n6) Engage state regulatory counsel",
        "Colorado: Feb 2026; Other states: various dates TBD",
        "Wealth managers operating across multiple states face growing compliance complexity. Colorado Act specifically covers high-risk AI in financial decisions.",
        "Colorado first comprehensive state AI law. Over 40 states introduced AI legislation in 2024-2025 legislative sessions.",
        "https://leg.colorado.gov/bills/sb24-205"
    ],
]

for r, row_data in enumerate(regulations, start_row + 1):
    for c, val in enumerate(row_data, 1):
        ws1.cell(row=r, column=c, value=val)
    style_row(ws1, r, len(headers), alt=(r % 2 == 0), center_cols=[1, 4, 5, 6, 11])

# Conditional formatting on Risk Level (column F)
risk_col = "F5:F100"
ws1.conditional_formatting.add(risk_col, CellIsRule(
    operator="equal", formula=['"Critical"'],
    fill=PatternFill("solid", fgColor="FEE2E2"), font=Font(color=CRITICAL_RED, bold=True)))
ws1.conditional_formatting.add(risk_col, CellIsRule(
    operator="equal", formula=['"High"'],
    fill=PatternFill("solid", fgColor="FFF7ED"), font=Font(color=HIGH_ORANGE, bold=True)))
ws1.conditional_formatting.add(risk_col, CellIsRule(
    operator="equal", formula=['"Medium"'],
    fill=PatternFill("solid", fgColor="FEF9C3"), font=Font(color=MEDIUM_YELLOW, bold=True)))
ws1.conditional_formatting.add(risk_col, CellIsRule(
    operator="equal", formula=['"Low"'],
    fill=PatternFill("solid", fgColor="DCFCE7"), font=Font(color=LOW_GREEN, bold=True)))

ws1.freeze_panes = "A5"
ws1.auto_filter.ref = f"A4:N{start_row + len(regulations)}"

# ============================================================
# SHEET 2 — Risk Summary Dashboard
# ============================================================
ws2 = wb.create_sheet("Risk Dashboard")
add_title(ws2, "Regulatory Risk Dashboard", "Summary metrics and risk distribution for compliance prioritization")

# KPI row
kpis = [
    ("Total Regulations", f"=COUNTA('Regulatory Register'!A5:A100)"),
    ("Critical Risk", f"=COUNTIF('Regulatory Register'!F5:F100,\"Critical\")"),
    ("High Risk", f"=COUNTIF('Regulatory Register'!F5:F100,\"High\")"),
    ("Medium Risk", f"=COUNTIF('Regulatory Register'!F5:F100,\"Medium\")"),
    ("Already Effective", f"=COUNTIFS('Regulatory Register'!E5:E100,\"*Effective*\")+COUNTIFS('Regulatory Register'!E5:E100,\"*Active*\")+COUNTIFS('Regulatory Register'!E5:E100,\"*Published*\")"),
]

for i, (label, formula) in enumerate(kpis):
    col = i + 1
    ws2.column_dimensions[get_column_letter(col)].width = 20
    ws2.cell(row=4, column=col, value=label)
    ws2.cell(row=4, column=col).font = Font(name="Arial", size=9, color="64748B")
    ws2.cell(row=4, column=col).alignment = Alignment(horizontal="center")
    ws2.cell(row=5, column=col, value=formula)
    ws2.cell(row=5, column=col).font = Font(name="Georgia", bold=True, size=22, color=NAVY)
    ws2.cell(row=5, column=col).alignment = Alignment(horizontal="center")
    for row in [4, 5]:
        ws2.cell(row=row, column=col).border = thin_border
ws2.row_dimensions[5].height = 40

# Impact area breakdown
ws2.cell(row=7, column=1, value="Impact Area Analysis")
ws2.cell(row=7, column=1).font = Font(name="Arial", bold=True, size=12, color=NAVY)

area_headers = ["Impact Area", "Regulation Count", "Critical Count", "Priority Action"]
for i, h in enumerate(area_headers, 1):
    ws2.cell(row=8, column=i, value=h)
    ws2.column_dimensions[get_column_letter(i)].width = [22, 18, 16, 55][i-1]
style_header(ws2, 8, 4)

areas = [
    ["Model Risk", "6", "3", "Expand AI model inventory; implement AI-specific validation under SR 11-7; deploy drift monitoring"],
    ["Disclosure / Conflicts", "5", "3", "Audit all AI marketing claims; update Form CRS; eliminate/neutralize AI conflicts of interest"],
    ["Fair Lending", "3", "2", "Conduct disparate impact testing on AI credit models; implement bias mitigation; document business necessity"],
    ["Consumer Protection", "5", "2", "Ensure Reg BI compliance for AI recommendations; validate adverse action explainability"],
    ["Data Privacy", "3", "0", "Map NPI in AI data pipelines; update privacy notices; implement consent management for AI processing"],
    ["Vendor Risk", "3", "0", "Tier AI vendors by criticality; conduct due diligence; review contracts; develop exit plans"],
]
for r, row_data in enumerate(areas, 9):
    for c, val in enumerate(row_data, 1):
        ws2.cell(row=r, column=c, value=val)
    style_row(ws2, r, 4, alt=(r % 2 == 0), center_cols=[2, 3])

# Regulatory body breakdown
ws2.cell(row=17, column=1, value="Regulatory Body Breakdown")
ws2.cell(row=17, column=1).font = Font(name="Arial", bold=True, size=12, color=NAVY)

body_headers = ["Regulator", "Guidance Count", "Key Focus"]
for i, h in enumerate(body_headers, 1):
    ws2.cell(row=18, column=i, value=h)
style_header(ws2, 18, 3)

bodies = [
    ["SEC", "4", "AI conflicts of interest, AI-washing enforcement, Marketing Rule, Reg BI, examination priorities"],
    ["FINRA", "2", "Supervisory obligations for AI tools, communications compliance, real-time AI monitoring"],
    ["Federal Reserve / OCC", "3", "Model risk management (SR 11-7), third-party risk, AI model validation"],
    ["CFPB", "2", "Adverse action explainability, fair lending/bias in AI credit models"],
    ["FTC / State", "2", "GLBA data privacy, CCPA/CPRA automated decision-making, state AI legislation"],
    ["NIST", "1", "AI Risk Management Framework — voluntary but increasingly referenced by regulators"],
    ["EU", "1", "EU AI Act — extraterritorial reach for US firms with EU clients or operations"],
]
for r, row_data in enumerate(bodies, 19):
    for c, val in enumerate(row_data, 1):
        ws2.cell(row=r, column=c, value=val)
    style_row(ws2, r, 3, alt=(r % 2 == 0), center_cols=[2])

# ============================================================
# SHEET 3 — Compliance Timeline
# ============================================================
ws3 = wb.create_sheet("Compliance Timeline")
add_title(ws3, "Compliance Implementation Timeline", "Phased deadlines and milestones for regulatory compliance actions")

t_headers = ["Phase", "Timeline", "Regulation(s)", "Key Actions", "Risk if Delayed", "Status"]
t_widths = [12, 18, 35, 55, 35, 14]

for i, (h, w) in enumerate(zip(t_headers, t_widths), 1):
    ws3.cell(row=4, column=i, value=h)
    ws3.column_dimensions[get_column_letter(i)].width = w
style_header(ws3, 4, len(t_headers))

timeline = [
    ["Phase 1", "Immediate (Now)", "SEC Marketing Rule; CFPB Circular 2022-03; Reg BI; SR 11-7; GLBA",
     "1) Audit AI marketing claims for substantiation\n2) Validate adverse action notice explainability\n3) Review AI recommendations for Reg BI compliance\n4) Expand model inventory to include all AI/ML\n5) Audit NPI handling in AI data pipelines",
     "Enforcement actions, fines, examination findings, reputational damage", "Urgent"],
    ["Phase 2", "Q3-Q4 2025", "FINRA RN 24-09; FINRA 2025 Report; SEC Exam Priorities; OCC 2023-17",
     "1) Implement real-time AI output monitoring\n2) Establish AI governance framework and AI CoE\n3) Conduct AI vendor due diligence and tiering\n4) Develop supervisor AI training programs\n5) Prepare AI examination documentation packages",
     "Examination deficiencies, MRAs, supervisory concerns", "In Progress"],
    ["Phase 3", "Q1-Q2 2026", "SEC PDA/AI Conflicts Rule (if finalized); NIST AI RMF adoption; State AI laws",
     "1) Implement conflict elimination/neutralization for AI tools\n2) Adopt NIST AI RMF Govern/Map/Measure/Manage framework\n3) Prepare for Colorado AI Act compliance (Feb 2026)\n4) Conduct enterprise-wide AI risk assessment\n5) Implement bias testing programs",
     "Non-compliance with new rules, state enforcement, competitive disadvantage", "Planning"],
    ["Phase 4", "H2 2026", "EU AI Act (high-risk full enforcement Aug 2026)",
     "1) Complete EU AI Act conformity assessments for high-risk AI\n2) Register AI systems with EU AI Office\n3) Implement quality management systems\n4) Ensure human oversight for all high-risk AI\n5) Deploy incident reporting mechanisms",
     "EU market access loss, penalties up to 7% global turnover", "Planning"],
    ["Ongoing", "Continuous", "All regulations — ongoing compliance monitoring",
     "1) Continuous AI model monitoring for drift and bias\n2) Annual model validation cycles\n3) Quarterly vendor risk reviews\n4) Regular fair lending testing\n5) Policy and procedure updates as regulations evolve\n6) Board and senior management reporting",
     "Compliance decay, emerging risk exposure", "Ongoing"],
]

for r, row_data in enumerate(timeline, 5):
    for c, val in enumerate(row_data, 1):
        ws3.cell(row=r, column=c, value=val)
    style_row(ws3, r, len(t_headers), alt=(r % 2 == 0), center_cols=[1, 2, 6])

ws3.conditional_formatting.add("F5:F100", CellIsRule(
    operator="equal", formula=['"Urgent"'],
    fill=PatternFill("solid", fgColor="FEE2E2"), font=Font(color=CRITICAL_RED, bold=True)))
ws3.conditional_formatting.add("F5:F100", CellIsRule(
    operator="equal", formula=['"In Progress"'],
    fill=PatternFill("solid", fgColor="FEF9C3"), font=Font(color=MEDIUM_YELLOW, bold=True)))
ws3.conditional_formatting.add("F5:F100", CellIsRule(
    operator="equal", formula=['"Planning"'],
    fill=PatternFill("solid", fgColor="DBEAFE"), font=Font(color=DARK_BLUE, bold=True)))
ws3.conditional_formatting.add("F5:F100", CellIsRule(
    operator="equal", formula=['"Ongoing"'],
    fill=PatternFill("solid", fgColor="DCFCE7"), font=Font(color=LOW_GREEN, bold=True)))

ws3.freeze_panes = "A5"

# ============================================================
# SHEET 4 — Enforcement Actions
# ============================================================
ws4 = wb.create_sheet("Enforcement Actions")
add_title(ws4, "AI-Related Enforcement Actions & Examination Findings",
          "Recent regulatory actions relevant to AI/ML in financial services")

e_headers = ["Date", "Agency", "Entity", "Action Type", "AI Issue", "Penalty / Outcome", "Relevance to WM", "Citation"]
e_widths = [12, 10, 22, 16, 40, 25, 40, 50]

for i, (h, w) in enumerate(zip(e_headers, e_widths), 1):
    ws4.cell(row=4, column=i, value=h)
    ws4.column_dimensions[get_column_letter(i)].width = w
style_header(ws4, 4, len(e_headers))

enforcements = [
    ["Mar 2024", "SEC", "Delphia (USA) Inc.", "Settlement",
     "False claims about using AI and ML to predict stock performance; claimed client data trained AI when technology did not exist",
     "$225,000 civil penalty",
     "Demonstrates SEC will enforce Marketing Rule against unsubstantiated AI claims by investment advisers",
     "SEC Release No. IA-6573 (Mar 18, 2024); https://www.sec.gov/newsroom/press-releases/2024-36"],
    ["Mar 2024", "SEC", "Global Predictions Inc.", "Settlement",
     "Misleading claims about AI/ML capabilities in investment process; touted AI-driven investment strategies without substance",
     "$175,000 civil penalty",
     "Confirms that even smaller advisers face scrutiny for AI-washing in marketing materials",
     "SEC Release No. IA-6574 (Mar 18, 2024); https://www.sec.gov/newsroom/press-releases/2024-36"],
    ["Oct 2024", "SEC", "Rimar Capital USA", "Enforcement",
     "Misrepresented AI trading capabilities; raised millions based on false claims of AI-automated multi-asset trading",
     "Pending / charges filed",
     "Escalation of AI-washing enforcement; demonstrates SEC willingness to pursue fraud charges for AI misrepresentation",
     "SEC enforcement action Oct 2024; https://www.dandodiary.com/2024/10/articles/artificial-intelligence/"],
    ["2024-2025", "CFPB", "Multiple institutions (unnamed)", "Examination findings",
     "AI credit scoring models produced disproportionately negative outcomes for Black and Hispanic applicants; insufficient CMS and testing for LDAs",
     "MRAs and supervisory actions",
     "Directly relevant to wealth managers offering credit products using AI underwriting models",
     "CFPB 2024 Fair Lending Report; https://www.consumerfinance.gov/data-research/research-reports/fair-lending-report-of-the-consumer-financial-protection-bureau-2024/"],
    ["2024-2025", "OCC/Fed", "Multiple banks (unnamed)", "Examination findings",
     "Incomplete AI model inventories; insufficient explainability documentation; weak ongoing monitoring for model drift and bias",
     "MRAs, supervisory concerns",
     "SR 11-7 deficiencies — examiners increasingly expecting AI-specific MRM enhancements at examined institutions",
     "Supervisory examination feedback; SR 11-7 implementation reviews"],
    ["Late 2023", "SEC", "Multiple advisers", "AI Sweep Examinations",
     "SEC Division of Examinations initiated broad information requests to advisers regarding AI use, capabilities, governance, and disclosures",
     "Information gathering / examinations",
     "Signals ongoing SEC focus on AI practices; results informing 2024-2025 enforcement strategy",
     "SEC Division of Examinations 2024-2025 priorities; https://www.sec.gov/about/divisions-offices/division-examinations/examination-priorities"],
]

for r, row_data in enumerate(enforcements, 5):
    for c, val in enumerate(row_data, 1):
        ws4.cell(row=r, column=c, value=val)
    style_row(ws4, r, len(e_headers), alt=(r % 2 == 0), center_cols=[1, 2, 4])

ws4.freeze_panes = "A5"
ws4.auto_filter.ref = f"A4:H{4 + len(enforcements)}"

# ============================================================
# SHEET 5 — 2025-2026 Outlook
# ============================================================
ws5 = wb.create_sheet("2025-2026 Outlook")
add_title(ws5, "2025-2026 Regulatory Outlook & Emerging Trends",
          "Forward-looking analysis of expected regulatory developments affecting AI in financial services")

o_headers = ["Area", "Expected Development", "Timeline", "Probability", "Impact", "Recommended Preparation"]
o_widths = [22, 50, 16, 14, 12, 55]

for i, (h, w) in enumerate(zip(o_headers, o_widths), 1):
    ws5.cell(row=4, column=i, value=h)
    ws5.column_dimensions[get_column_letter(i)].width = w
style_header(ws5, 4, len(o_headers))

outlook = [
    ["SEC PDA/AI Rule", "Finalization of SEC proposed rule on conflicts of interest from predictive data analytics. May be narrowed from original proposal scope under current administration but core principles likely retained.", "2025-2026", "Medium", "Critical",
     "Monitor rulemaking progress; prepare conflict mapping framework for AI tools; develop elimination/neutralization procedures regardless of final rule scope"],
    ["CFPB Reg B Amendments", "Proposed elimination of disparate impact liability under ECOA. Even if finalized, state laws and FHA retain disparate impact claims. Firms should maintain robust bias testing.", "2025-2026", "Medium-High", "High",
     "Continue disparate impact testing despite proposed narrowing; maintain documentation; monitor state fair lending developments; do not relax bias mitigation programs"],
    ["GLBA Modernization", "Congressional discussion draft to modernize GLBA with stronger data minimization, AI transparency requirements, and potential preemption of state laws.", "2026-2027", "Medium", "High",
     "Monitor legislative progress; prepare data minimization capabilities; document AI data usage transparently; engage industry trade groups"],
    ["EU AI Act Full Enforcement", "High-risk AI system requirements become fully enforceable in August 2026. US firms serving EU clients must achieve conformity.", "Aug 2026", "Certain", "High",
     "Begin conformity assessments now; implement quality management systems; prepare technical documentation; budget for compliance investment"],
    ["State AI Legislation Wave", "Multiple states expected to pass AI regulation following Colorado's lead. May create patchwork compliance landscape rivaling state privacy law complexity.", "2025-2027", "High", "Medium",
     "Monitor legislation in key operating states; build flexible AI governance adaptable to state requirements; consider hiring state regulatory specialists"],
    ["NIST AI RMF Adoption", "Expected to become de facto standard referenced in federal examinations. Updated GenAI profiles and threat taxonomies expanding framework scope.", "Ongoing", "High", "Medium",
     "Adopt NIST AI RMF as baseline framework; integrate with SR 11-7 MRM programs; use FS AI RMF for sector-specific implementation"],
    ["AI-Specific Banking Guidance", "Federal banking agencies (OCC/Fed/FDIC) expected to issue updated AI-specific guidance supplementing SR 11-7/OCC 2011-12 for modern AI/ML systems.", "2025-2026", "Medium-High", "Critical",
     "Maintain strong SR 11-7 compliance as foundation; enhance AI model validation procedures; prepare for potentially prescriptive AI governance requirements"],
    ["Interagency AI Coordination", "Increased coordination between SEC, FINRA, CFPB, OCC, and state regulators on AI oversight. Potential for harmonized expectations or joint enforcement.", "2025-2026", "Medium", "High",
     "Prepare for multi-regulator examination scenarios; ensure consistent AI governance documentation across regulatory requirements; build unified compliance program"],
]

for r, row_data in enumerate(outlook, 5):
    for c, val in enumerate(row_data, 1):
        ws5.cell(row=r, column=c, value=val)
    style_row(ws5, r, len(o_headers), alt=(r % 2 == 0), center_cols=[3, 4, 5])

ws5.conditional_formatting.add("E5:E100", CellIsRule(
    operator="equal", formula=['"Critical"'],
    fill=PatternFill("solid", fgColor="FEE2E2"), font=Font(color=CRITICAL_RED, bold=True)))
ws5.conditional_formatting.add("E5:E100", CellIsRule(
    operator="equal", formula=['"High"'],
    fill=PatternFill("solid", fgColor="FFF7ED"), font=Font(color=HIGH_ORANGE, bold=True)))
ws5.conditional_formatting.add("E5:E100", CellIsRule(
    operator="equal", formula=['"Medium"'],
    fill=PatternFill("solid", fgColor="FEF9C3"), font=Font(color=MEDIUM_YELLOW, bold=True)))

ws5.freeze_panes = "A5"

# --- Save ---
output = "ai-ml-regulatory-compliance-landscape-2025-2026.xlsx"
wb.save(output)
size_kb = os.path.getsize(output) / 1024
print(f"Created: {output} ({size_kb:.1f} KB)")
print(f"Sheets: {', '.join(wb.sheetnames)}")
