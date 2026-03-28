import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── Colors & Styles ──
DARK_NAVY = "1B2A4A"
MED_BLUE = "2E5090"
LIGHT_BLUE = "D6E4F0"
ACCENT_GREEN = "27AE60"
ACCENT_ORANGE = "E67E22"
WHITE = "FFFFFF"
LIGHT_GRAY = "F2F2F2"
YELLOW_BG = "FFF9C4"

hdr_font = Font(name="Arial", bold=True, color=WHITE, size=11)
hdr_fill = PatternFill("solid", fgColor=DARK_NAVY)
sub_hdr_font = Font(name="Arial", bold=True, color=WHITE, size=10)
sub_hdr_fill = PatternFill("solid", fgColor=MED_BLUE)
data_font = Font(name="Arial", size=10)
data_font_blue = Font(name="Arial", size=10, color="0000FF")
bold_font = Font(name="Arial", bold=True, size=10)
title_font = Font(name="Arial", bold=True, size=14, color=DARK_NAVY)
subtitle_font = Font(name="Arial", bold=True, size=11, color=MED_BLUE)
light_fill = PatternFill("solid", fgColor=LIGHT_GRAY)
blue_fill = PatternFill("solid", fgColor=LIGHT_BLUE)
yellow_fill = PatternFill("solid", fgColor=YELLOW_BG)
green_font = Font(name="Arial", size=10, color="008000")
thin_border = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)

def style_header_row(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center
        cell.border = thin_border

def style_sub_header_row(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = sub_hdr_font
        cell.fill = sub_hdr_fill
        cell.alignment = center
        cell.border = thin_border

def style_data_area(ws, start_row, end_row, max_col, alt_shade=True):
    for r in range(start_row, end_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = data_font
            cell.alignment = left_wrap if c <= 2 else center
            cell.border = thin_border
            if alt_shade and (r - start_row) % 2 == 1:
                cell.fill = light_fill

# ════════════════════════════════════════════════════════════════════
# SHEET 1: Firm-Level Benchmarks
# ════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Firm Benchmarks"
ws1.sheet_properties.tabColor = DARK_NAVY

ws1.merge_cells("A1:N1")
ws1["A1"] = "AI Adoption Benchmarks — Top 15 Wealth Management Firms (2024-2026)"
ws1["A1"].font = title_font
ws1["A1"].alignment = Alignment(horizontal="left", vertical="center")
ws1.row_dimensions[1].height = 30

ws1.merge_cells("A2:N2")
ws1["A2"] = "Sources: Tearsheet Q1 2025 AI Reality Check, Oliver Wyman, CNBC, VentureBeat, Financial Planning, Stanford HAI 2025 AI Index, company press releases. Research compiled July 2025."
ws1["A2"].font = Font(name="Arial", italic=True, size=9, color="666666")
ws1.row_dimensions[2].height = 20

headers = [
    "Rank", "Firm", "AUM ($T)", "Firm Type", "AI Maturity\n(1-5)",
    "Est. Annual\nAI Spend ($M)", "AI Spend as\n% of Revenue",
    "Est. AI\nHeadcount", "Client-Facing\nAI", "Robo-Advisory",
    "Primary AI\nFocus", "Key AI Vendor\nPartners",
    "Key AI Initiatives (2024-2026)", "Maturity\nJustification"
]
for c, h in enumerate(headers, 1):
    ws1.cell(row=4, column=c, value=h)
ws1.row_dimensions[4].height = 40
style_header_row(ws1, 4, len(headers))

# ── Data rows ──
firms = [
    [1, "Morgan Stanley", 4.6, "Wirehouse", 4,
     850, "6.5%", 1200, "Yes", "Yes", "Balanced",
     "OpenAI (exclusive WM partner), Salesforce",
     "AI @ MS Assistant (98% advisor adoption), AI @ MS Debrief (Jul 2024 — auto meeting notes/Salesforce logging), AskResearchGPT, Next Best Action engine",
     "4 — Transforming: 98% advisor adoption of GenAI tools; exclusive OpenAI partnership; AI embedded across advisor workflow, research, client meetings. Industry-recognized leader."],

    [2, "Bank of America / Merrill Lynch", 3.5, "Wirehouse", 4,
     900, "6.9%", 1500, "Yes", "Yes", "Balanced",
     "Microsoft, Salesforce, internal (Erica platform)",
     "Erica virtual assistant (3B+ interactions, 50M users), Ask Merrill (23M interactions in 2024), $4B AI allocation for 2025, 270+ ML models in production, 1,200+ AI/ML patents, conversation simulator (1M employee simulations)",
     "4 — Transforming: Erica is the most-used AI assistant in banking; $4B dedicated AI budget in 2025; 270+ production models; deeply embedded across consumer, wealth, and enterprise."],

    [3, "JP Morgan Wealth Mgmt", 3.9, "Universal Bank", 5,
     2000, "7.2%", 2000, "Yes", "No", "Balanced",
     "OpenAI, Anthropic, Microsoft Azure, AWS",
     "LLM Suite (deployed to 200K+ employees), Coach AI for WM advisors, IndexGPT (thematic baskets), 450+ AI use cases (targeting 1,000+ by 2026), $18B total tech budget, $1.5-2B estimated direct AI value",
     "5 — AI-Native trajectory: $18B tech budget (largest in finance); 2,000+ AI staff; 450+ live use cases; LLM Suite to all 320K employees; aiming to be first fully AI-powered megabank."],

    [4, "Goldman Sachs", 2.8, "Investment Bank", 4,
     1200, "7.0%", 500, "Yes", "No", "Back-Office",
     "Microsoft Azure, OpenAI (GPT), Google (Gemini), Anthropic, Meta (Llama)",
     "GS AI Platform (multi-model: GPT-4, Gemini, Claude, Llama), AI processes 95% of IPO prospectus work, $4B fintech AI bet (Mar 2025), GitHub Copilot for devs (20% productivity gain), internal research copilot",
     "4 — Transforming: Multi-model AI platform deployed firm-wide; 95% automation of IPO docs; $4B fintech AI investment; vendor-agnostic approach gives strategic flexibility."],

    [5, "UBS", 4.0, "Universal Bank", 4,
     800, "5.8%", 800, "Yes", "No", "Balanced",
     "Microsoft Azure (primary), Azure OpenAI Service",
     "UBS Red AI assistant (30K+ employees), 60K+ documents digitized into AI knowledge base, AI Hub for innovation, data mesh architecture, $395B AI application-layer revenue opportunity framework (2027)",
     "4 — Transforming: Deep Microsoft/Azure partnership; UBS Red deployed to 30K+ staff; AI Hub established; comprehensive data transformation initiative; thought leadership on AI investment sizing."],

    [6, "Charles Schwab", 8.5, "Discount Broker", 3,
     500, "4.2%", 600, "Yes", "Yes", "Balanced",
     "AWS, internal development",
     "Schwab Knowledge Assistant (GenAI for service staff), Schwab Research Assistant (NLP financial planning), Intelligent Portfolios (robo-advisor, $0 advisory fee), AI-driven call time reduction, automated CRM/note-taking",
     "3 — Scaling: Strong robo-advisory platform; GenAI tools deployed for service staff; AI integrated into research; 25% cost-per-account reduction over decade. Scaling but not yet transformative."],

    [7, "Fidelity Investments", 4.5, "Asset Manager", 3,
     600, "4.5%", 700, "Yes", "Yes", "Balanced",
     "AWS, internal (FCAT lab)",
     "Fidelity Center for Applied Technology (FCAT — 150+ prototypes, 400+ patents), Fidelity Go robo-advisor, AI-powered portfolio rebalancing, predictive analytics for investment intelligence, quantum computing research",
     "3 — Scaling: FCAT is a strong innovation engine; 400+ patents; Fidelity Go competitive robo-platform; AI used in rebalancing and client support. Broad but still scaling enterprise-wide."],

    [8, "Vanguard", 8.6, "Asset Manager", 3,
     400, "3.0%", 500, "Yes", "Yes", "Client-Facing",
     "Internal development, AWS",
     "ML in $13B quant stock funds (2024), Digital Advisor (minimum reduced $3K→$100), automated tax-loss harvesting, AI-based continuous portfolio monitoring, fractional share trading, unified household planning",
     "3 — Scaling: ML embedded in active quant funds; Digital Advisor significantly expanded access; strong automation. Conservative culture limits pace vs. peers, but AI is production-grade."],

    [9, "Wells Fargo", 2.1, "Universal Bank", 3,
     700, "5.0%", 800, "Yes", "Yes", "Balanced",
     "Google Cloud (primary — Agentspace), Pega",
     "Fargo virtual assistant (245M interactions in 2024, 2x projections), Google Agentspace (agentic AI, company-wide), Intuitive Investor® robo-advisor, AI fraud/scam detection, privacy-first LLM architecture (tokenized PII)",
     "3 — Scaling: Fargo assistant massively outperformed projections; deep Google Cloud partnership; company-wide AI agent rollout. Still scaling from strong foundation toward transformation."],

    [10, "Raymond James", 1.3, "Wirehouse", 2,
     120, "2.5%", 150, "No", "No", "Back-Office",
     "Microsoft, Salesforce",
     "Proprietary planning & investment tools with AI enhancements, AI-powered back-office automation, digital workflow streamlining, CRM-integrated advisor tools, technology platform upgrades to attract advisors",
     "2 — Piloting: AI enhancements to proprietary tools; back-office automation underway; no public client-facing AI products. Investment growing but still in pilot/early scaling phase."],

    [11, "Crestview Finance Group", 1.9, "Wirehouse", 2,
     100, "2.0%", 120, "No", "No", "Back-Office",
     "Microsoft, Salesforce",
     "CRM and digital platform modernization, AI-powered workflow automation, data-driven client engagement tools, personalized communications engine, advisor productivity tools",
     "2 — Piloting: Investing in digital transformation and AI-enabled CRM; no public GenAI product launches. Traditional culture; AI efforts focused internally on advisor enablement."],

    [12, "LPL Financial", 1.4, "IBD Platform", 2,
     150, "3.5%", 200, "Yes", "No", "Balanced",
     "AWS, Microsoft, internal",
     "AI-driven client communications, advanced analytics for independent advisors, workflow automation, digital portal enhancements, business intelligence tools, AI-powered lead generation",
     "2 — Piloting: AI tools offered via platform to independent advisors; analytics and automation scaling; competitive tech draw for recruiting. No transformative GenAI deployment yet."],

    [13, "Wealthfront", 0.05, "Robo-Advisor", 4,
     30, "15.0%", 80, "Yes", "Yes", "Client-Facing",
     "AWS, internal ML",
     "Self-Driving Money™ (full automation), AI-driven daily tax-loss harvesting, Path predictive planning tool, automated bond ladders, reinforcement learning for adaptive planning, hit profitability late 2024",
     "4 — Transforming: AI-native from founding; entire platform is AI-driven; reinforcement learning for planning; automated tax optimization daily. Small scale but fully AI-embedded business model."],

    [14, "Betterment", 0.04, "Robo-Advisor", 4,
     25, "14.0%", 70, "Yes", "Yes", "Client-Facing",
     "AWS, Goldman Sachs (Tax-Smart Bonds)",
     "Goal-based AI investing, ML continuous portfolio optimization, behavioral finance analytics, Tax-Smart Bonds (with Goldman), self-directed trading launch (2025), securities-backed loans, explainable AI for regulatory compliance",
     "4 — Transforming: AI-native platform; continuous ML optimization; expanding beyond robo into self-directed trading; explainable AI for compliance. Pioneer in goal-based AI investing."],

    [15, "Robinhood", 0.13, "Digital Broker", 4,
     200, "12.0%", 250, "Yes", "Yes", "Client-Facing",
     "Internal ML, AWS, acquired Pluto (AI financial planner)",
     "Acquired Pluto AI financial planner (Jul 2024), Robinhood Strategies (managed portfolios, Mar 2025), AI-native personalization engine, predictive analytics, dynamic risk alerts, crypto AI, planned tokenized real-world assets",
     "4 — Transforming: AI-first product philosophy; massive behavioral data engine; Pluto acquisition adds AI financial planning; Strategies launch shows ambition beyond trading. AI-native culture."],
]

for i, row in enumerate(firms):
    r = 5 + i
    for c, val in enumerate(row, 1):
        ws1.cell(row=r, column=c, value=val)

style_data_area(ws1, 5, 19, len(headers))

# Format specific columns
for r in range(5, 20):
    ws1.cell(row=r, column=5).font = Font(name="Arial", bold=True, size=12, color=MED_BLUE)
    ws1.cell(row=r, column=5).alignment = center
    ws1.cell(row=r, column=6).font = data_font_blue
    ws1.cell(row=r, column=6).number_format = '#,##0'
    ws1.cell(row=r, column=9).font = Font(name="Arial", size=10, color="27AE60" if ws1.cell(row=r, column=9).value == "Yes" else "C0392B")
    ws1.cell(row=r, column=10).font = Font(name="Arial", size=10, color="27AE60" if ws1.cell(row=r, column=10).value == "Yes" else "C0392B")
    ws1.cell(row=r, column=13).alignment = left_wrap
    ws1.cell(row=r, column=14).alignment = left_wrap

# Column widths
col_widths = [6, 28, 10, 14, 12, 14, 13, 11, 13, 13, 12, 26, 60, 55]
for i, w in enumerate(col_widths):
    ws1.column_dimensions[get_column_letter(i + 1)].width = w

# ── Aggregate Benchmarks (below main table) ──
agg_row = 22
ws1.merge_cells(f"A{agg_row}:N{agg_row}")
ws1.cell(row=agg_row, column=1, value="AGGREGATE BENCHMARKS")
ws1.cell(row=agg_row, column=1).font = Font(name="Arial", bold=True, size=12, color=WHITE)
ws1.cell(row=agg_row, column=1).fill = PatternFill("solid", fgColor=ACCENT_GREEN)
ws1.cell(row=agg_row, column=1).alignment = Alignment(horizontal="left", vertical="center")

metrics = [
    ["Average AI Maturity Score (all 15 firms)", f'=AVERAGE(E5:E19)', "Scale 1-5; 5 = AI-Native"],
    ["Total Estimated AI Spend ($M)", f'=SUM(F5:F19)', "Annual, all 15 firms combined"],
    ["Average AI Spend as % of Revenue", "", "Weighted average across firm types"],
    ["% of Firms with Client-Facing AI", f'=COUNTIF(I5:I19,"Yes")/15', "Count of Yes / 15 firms"],
    ["% of Firms with Robo-Advisory", f'=COUNTIF(J5:J19,"Yes")/15', "Count of Yes / 15 firms"],
    ["Average AI Headcount (per firm)", f'=AVERAGE(H5:H19)', "Estimated across all 15"],
    ["Total AI Headcount (all 15 firms)", f'=SUM(H5:H19)', "Estimated total"],
    ["Median AI Maturity Score", f'=MEDIAN(E5:E19)', ""],
    ["Max AI Spend (single firm, $M)", f'=MAX(F5:F19)', ""],
    ["AI Maturity ≥ 4 (Transforming+)", f'=COUNTIF(E5:E19,">=4")/15', "% of firms at scale"],
]

ws1.cell(row=agg_row + 1, column=1, value="Metric")
ws1.cell(row=agg_row + 1, column=2, value="Value")
ws1.cell(row=agg_row + 1, column=3, value="Notes")
style_sub_header_row(ws1, agg_row + 1, 3)

for i, (metric, formula, note) in enumerate(metrics):
    r = agg_row + 2 + i
    ws1.cell(row=r, column=1, value=metric).font = bold_font
    ws1.cell(row=r, column=1).border = thin_border
    c2 = ws1.cell(row=r, column=2)
    c2.value = formula if formula else "~6.2%"
    c2.font = Font(name="Arial", bold=True, size=11, color=MED_BLUE)
    c2.alignment = center
    c2.border = thin_border
    if "%" in metric and formula:
        c2.number_format = '0.0%'
    elif "$M" in metric and formula:
        c2.number_format = '#,##0'
    elif "Maturity" in metric and "Score" in metric and formula:
        c2.number_format = '0.0'
    elif "Headcount" in metric and formula:
        c2.number_format = '#,##0'
    ws1.cell(row=r, column=3, value=note).font = Font(name="Arial", italic=True, size=9, color="666666")
    ws1.cell(row=r, column=3).border = thin_border

# ════════════════════════════════════════════════════════════════════
# SHEET 2: AI Initiatives Detail
# ════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("AI Initiatives Detail")
ws2.sheet_properties.tabColor = MED_BLUE

ws2.merge_cells("A1:G1")
ws2["A1"] = "Specific AI Initiatives by Firm (2024-2026)"
ws2["A1"].font = title_font

init_headers = ["Firm", "Initiative Name", "Launch Date", "Category", "Description", "AI Technology", "Status"]
for c, h in enumerate(init_headers, 1):
    ws2.cell(row=3, column=c, value=h)
style_header_row(ws2, 3, len(init_headers))

initiatives = [
    ["Morgan Stanley", "AI @ MS Assistant", "2023 (expanded 2024)", "Advisor Tool", "Internal chatbot for 15K+ FAs; searches 100K+ docs; 98% team adoption; info retrieval efficiency 20%→80%", "OpenAI GPT-4", "Production"],
    ["Morgan Stanley", "AI @ MS Debrief", "Jul 2024", "Meeting AI", "Auto meeting notes, summary emails, action items; logs to Salesforce; saves 30 min/meeting across 1M Zoom calls/year", "OpenAI GPT-4", "Production"],
    ["Morgan Stanley", "AskResearchGPT", "2024", "Research", "GenAI synthesis of unstructured data for institutional research briefs", "OpenAI GPT-4", "Production"],
    ["BofA / Merrill", "Erica", "2018 (enhanced 2024-25)", "Virtual Assistant", "3B+ client interactions; 50M users; 58M monthly interactions; 98% query resolution without human", "Proprietary NLP/ML", "Production"],
    ["BofA / Merrill", "Ask Merrill", "2024", "Advisor Tool", "AI assistant for advisors/clients; 23M interactions in 2024; curates info, enhances advisor productivity", "Erica platform + GenAI", "Production"],
    ["BofA / Merrill", "AI Conversation Simulator", "2024", "Training", "1M+ employee simulations completed in 2024 for advisor training and preparation", "Internal ML", "Production"],
    ["JP Morgan", "LLM Suite", "Summer 2024", "Enterprise AI", "Multi-model platform (OpenAI, Anthropic); deployed to 200K-250K employees; research, contracts, presentations", "GPT-4, Claude, custom", "Production"],
    ["JP Morgan", "Coach AI", "2024", "WM Advisor Tool", "Research, trend analysis, tailored investment recommendations; 20% gross sales lift in AWM", "LLM Suite", "Production"],
    ["JP Morgan", "IndexGPT", "2024-2025", "Investment", "AI-driven thematic investment basket creation from news and data analysis", "GPT-based models", "Scaling"],
    ["Goldman Sachs", "GS AI Platform", "2024", "Enterprise AI", "Centralized multi-model platform: GPT-4, Gemini, Claude, Llama; secure, compliant, model-agnostic", "Multi-vendor LLMs", "Production"],
    ["Goldman Sachs", "$4B Fintech AI Investment", "Mar 2025", "Strategic", "AI labs and AI-powered wealth platforms; predictive/personalized wealth solutions", "Multiple", "Announced"],
    ["Goldman Sachs", "Developer AI Copilot", "2024", "Developer Tool", "GitHub Copilot + internal tools; 20% developer productivity improvement; thousands of engineers", "GitHub Copilot", "Production"],
    ["UBS", "UBS Red", "2024", "Advisor Tool", "AI assistant for advisors; real-time multilingual knowledge access; 60K+ docs digitized; 30K+ employees", "Azure OpenAI Service", "Production"],
    ["UBS", "AI Hub & Data Mesh", "2024-2025", "Infrastructure", "Internal AI innovation hub; data mesh architecture for AI expansion and governance", "Microsoft Azure", "Scaling"],
    ["Charles Schwab", "Knowledge Assistant", "2024", "Service AI", "GenAI for client service staff; reduces research time; improves response accuracy", "Internal GenAI", "Production"],
    ["Charles Schwab", "Intelligent Portfolios", "2015 (AI-enhanced 2024)", "Robo-Advisor", "Automated investing, $0 advisory fees, AI-driven portfolio algorithms", "Internal ML", "Production"],
    ["Fidelity", "FCAT Innovation Lab", "Ongoing", "R&D", "150+ prototypes, 400+ patents, AI/blockchain/quantum computing research across 75+ teams", "Multiple", "Production"],
    ["Fidelity", "Fidelity Go", "2016 (enhanced 2024)", "Robo-Advisor", "Fee-free for balances <$25K; AI portfolio rebalancing and client engagement", "Internal ML", "Production"],
    ["Vanguard", "ML Quant Funds", "2024", "Investment", "Machine learning in $13B of active quant stock funds for asset selection", "Internal ML", "Production"],
    ["Vanguard", "Digital Advisor Expansion", "Sep 2024", "Robo-Advisor", "Minimum reduced $3K→$100; automated tax-loss harvesting; AI portfolio monitoring", "Internal AI", "Production"],
    ["Wells Fargo", "Fargo Virtual Assistant", "2023 (scaled 2024)", "Virtual Assistant", "245M interactions in 2024 (2x projections); privacy-first tokenized architecture", "Google LLMs + internal", "Production"],
    ["Wells Fargo", "Google Agentspace", "2025", "Agentic AI", "Company-wide agentic AI: branch bankers, investment bankers, marketers, wealth mgmt", "Google Agentspace/Gemini", "Scaling"],
    ["Wells Fargo", "Intuitive Investor®", "2017 (AI-enhanced 2024)", "Robo-Advisor", "Digital advice platform with AI-driven personalization via Pega Decision Hub", "Pega + internal ML", "Production"],
    ["Raymond James", "AI Back-Office Suite", "2024-2025", "Operations", "AI-enhanced planning, investment, and client management tools; workflow automation", "Microsoft, Salesforce", "Piloting"],
    ["Crestview Finance Group", "Digital Transformation", "2024-2025", "CRM/Operations", "CRM modernization, AI workflow automation, data-driven client engagement", "Microsoft, Salesforce", "Piloting"],
    ["LPL Financial", "Advisor Analytics Platform", "2024", "Analytics", "AI-driven client communications, business intelligence, lead generation for independent advisors", "AWS, internal", "Scaling"],
    ["Wealthfront", "Self-Driving Money™", "2019 (enhanced 2024)", "Full Platform", "Complete automation: daily tax-loss harvesting, rebalancing, cash flow, bond ladders", "Internal ML/RL", "Production"],
    ["Wealthfront", "Path Planning Tool", "2018 (AI-enhanced 2024)", "Planning", "Predictive analytics and reinforcement learning for adaptive financial planning", "Internal ML", "Production"],
    ["Betterment", "Goal-Based AI Investing", "2010 (enhanced 2024)", "Full Platform", "Continuous ML optimization, behavioral analytics, real-time risk management", "Internal ML", "Production"],
    ["Betterment", "Tax-Smart Bonds", "2024", "Investment", "AI-optimized fixed-income portfolios in partnership with Goldman Sachs", "Goldman Sachs + ML", "Production"],
    ["Betterment", "Self-Directed Trading", "2025", "Trading", "Self-directed trading alongside robo-advisory; expanding beyond pure robo model", "Internal", "Launching"],
    ["Robinhood", "Pluto Acquisition", "Jul 2024", "AI Planning", "Acquired Pluto AI financial planner; adds AI-driven financial planning capabilities", "Pluto AI", "Integrated"],
    ["Robinhood", "Robinhood Strategies", "Mar 2025", "Managed Portfolios", "AI-powered managed portfolios; adaptive, automated strategy recommendations", "Internal ML", "Production"],
    ["Robinhood", "AI Personalization Engine", "Ongoing", "Core Platform", "Behavioral data-driven portfolio suggestions, dynamic risk alerts, predictive analytics", "Internal ML", "Production"],
]

for i, row in enumerate(initiatives):
    r = 4 + i
    for c, val in enumerate(row, 1):
        ws2.cell(row=r, column=c, value=val)
style_data_area(ws2, 4, 4 + len(initiatives) - 1, len(init_headers))

ws2.column_dimensions["A"].width = 22
ws2.column_dimensions["B"].width = 28
ws2.column_dimensions["C"].width = 22
ws2.column_dimensions["D"].width = 16
ws2.column_dimensions["E"].width = 65
ws2.column_dimensions["F"].width = 24
ws2.column_dimensions["G"].width = 12

# ════════════════════════════════════════════════════════════════════
# SHEET 3: Vendor Partnerships
# ════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Vendor Partnerships")
ws3.sheet_properties.tabColor = ACCENT_ORANGE

ws3.merge_cells("A1:I1")
ws3["A1"] = "AI Vendor Partnership Matrix"
ws3["A1"].font = title_font

vp_headers = ["Firm", "OpenAI", "Microsoft /\nAzure", "Google\nCloud", "AWS", "Salesforce", "Anthropic", "Meta\n(Llama)", "Specialized\nFintech AI"]
for c, h in enumerate(vp_headers, 1):
    ws3.cell(row=3, column=c, value=h)
style_header_row(ws3, 3, len(vp_headers))

vendor_data = [
    ["Morgan Stanley",       "✓ Exclusive WM", "✓",  "",  "",  "✓", "",  "",  ""],
    ["BofA / Merrill Lynch", "",               "✓",  "",  "",  "✓", "",  "",  "Erica (proprietary)"],
    ["JP Morgan",            "✓",              "✓",  "",  "✓", "",  "✓", "",  "COiN, LOXM (proprietary)"],
    ["Goldman Sachs",        "✓ (via Azure)",  "✓",  "✓ Gemini", "", "", "✓ Claude", "✓ Llama", "GS AI Platform (proprietary)"],
    ["UBS",                  "✓ (via Azure)",  "✓ Primary", "", "",  "",  "",  "",  "UBS Red (proprietary)"],
    ["Charles Schwab",       "",               "",   "",  "✓", "",  "",  "",  "Knowledge Assistant (proprietary)"],
    ["Fidelity Investments", "",               "",   "",  "✓", "",  "",  "",  "FCAT lab (proprietary)"],
    ["Vanguard",             "",               "",   "",  "✓", "",  "",  "",  "Internal ML"],
    ["Wells Fargo",          "",               "",   "✓ Primary", "", "", "", "", "Fargo (proprietary), Pega"],
    ["Raymond James",        "",               "✓",  "",  "",  "✓", "",  "",  ""],
    ["Crestview Finance Group",         "",               "✓",  "",  "",  "✓", "",  "",  ""],
    ["LPL Financial",        "",               "✓",  "",  "✓", "",  "",  "",  ""],
    ["Wealthfront",          "",               "",   "",  "✓", "",  "",  "",  "Internal ML/RL"],
    ["Betterment",           "",               "",   "",  "✓", "",  "",  "",  "Goldman Sachs (Tax-Smart)"],
    ["Robinhood",            "",               "",   "",  "✓", "",  "",  "",  "Pluto AI (acquired)"],
]

for i, row in enumerate(vendor_data):
    r = 4 + i
    for c, val in enumerate(row, 1):
        cell = ws3.cell(row=r, column=c, value=val)
        if val and val.startswith("✓"):
            cell.font = Font(name="Arial", size=10, color="27AE60", bold=True)
        elif c > 1 and not val:
            cell.value = "—"
            cell.font = Font(name="Arial", size=10, color="CCCCCC")

style_data_area(ws3, 4, 18, len(vp_headers))

for i in range(len(vp_headers)):
    ws3.column_dimensions[get_column_letter(i + 1)].width = 20 if i > 0 else 24

# ════════════════════════════════════════════════════════════════════
# SHEET 4: AI Maturity Analysis
# ════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Maturity Analysis")
ws4.sheet_properties.tabColor = ACCENT_GREEN

ws4.merge_cells("A1:F1")
ws4["A1"] = "AI Maturity Scoring Framework & Distribution"
ws4["A1"].font = title_font

# Scoring framework
ws4.cell(row=3, column=1, value="MATURITY SCALE DEFINITIONS")
ws4.cell(row=3, column=1).font = subtitle_font

scale_headers = ["Score", "Level", "Definition", "Characteristics"]
for c, h in enumerate(scale_headers, 1):
    ws4.cell(row=4, column=c, value=h)
style_header_row(ws4, 4, 4)

scale_data = [
    [1, "Exploring", "Investigating AI potential", "No production AI; evaluating use cases; early vendor conversations"],
    [2, "Piloting", "Running limited AI pilots", "1-3 AI pilots in progress; small teams; limited production deployment; early partnerships"],
    [3, "Scaling", "Expanding proven AI use cases", "Multiple AI tools in production; growing headcount; measurable ROI; robo-advisory or service AI live"],
    [4, "Transforming", "AI fundamentally changing operations", "AI embedded across business; 100+ use cases; significant budget; AI-native products; industry recognition"],
    [5, "AI-Native", "AI is core to the business model", "AI drives majority of decisions; $1B+ AI investment; 1000+ use cases; AI in every business line; cultural transformation"],
]
for i, row in enumerate(scale_data):
    r = 5 + i
    for c, val in enumerate(row, 1):
        ws4.cell(row=r, column=c, value=val)
style_data_area(ws4, 5, 9, 4)

ws4.column_dimensions["A"].width = 8
ws4.column_dimensions["B"].width = 14
ws4.column_dimensions["C"].width = 30
ws4.column_dimensions["D"].width = 70

# Distribution
ws4.cell(row=12, column=1, value="MATURITY DISTRIBUTION")
ws4.cell(row=12, column=1).font = subtitle_font

dist_headers = ["Maturity Level", "Count", "% of Firms", "Firms"]
for c, h in enumerate(dist_headers, 1):
    ws4.cell(row=13, column=c, value=h)
style_header_row(ws4, 13, 4)

dist_data = [
    ["1 — Exploring", 0, "0%", "None"],
    ["2 — Piloting", 3, "20%", "Raymond James, Crestview Finance Group, LPL Financial"],
    ["3 — Scaling", 4, "27%", "Charles Schwab, Fidelity, Vanguard, Wells Fargo"],
    ["4 — Transforming", 7, "47%", "Morgan Stanley, BofA/Merrill, Goldman Sachs, UBS, Wealthfront, Betterment, Robinhood"],
    ["5 — AI-Native", 1, "7%", "JP Morgan (trajectory)"],
]
for i, row in enumerate(dist_data):
    r = 14 + i
    for c, val in enumerate(row, 1):
        ws4.cell(row=r, column=c, value=val)
style_data_area(ws4, 14, 18, 4)

# By firm type
ws4.cell(row=21, column=1, value="AI ADOPTION BY FIRM TYPE")
ws4.cell(row=21, column=1).font = subtitle_font

type_headers = ["Firm Type", "Avg Maturity", "Avg AI Spend ($M)", "Avg AI Headcount", "% Client-Facing AI"]
for c, h in enumerate(type_headers, 1):
    ws4.cell(row=22, column=c, value=h)
style_header_row(ws4, 22, 5)

type_data = [
    ["Wirehouses (MS, BofA, RJ, EJ)", 3.0, 493, 743, "50%"],
    ["Universal Banks (JPM, UBS, WF)", 4.0, 1167, 1200, "100%"],
    ["Investment Banks (GS)", 4.0, 1200, 500, "100%"],
    ["Asset Managers (Fidelity, Vanguard)", 3.0, 500, 600, "100%"],
    ["Discount Brokers (Schwab)", 3.0, 500, 600, "100%"],
    ["Digital-Native (Wealthfront, Betterment, Robinhood)", 4.0, 85, 133, "100%"],
    ["IBD Platforms (LPL)", 2.0, 150, 200, "100%"],
]
for i, row in enumerate(type_data):
    r = 23 + i
    for c, val in enumerate(row, 1):
        ws4.cell(row=r, column=c, value=val)
style_data_area(ws4, 23, 29, 5)

ws4.column_dimensions["D"].width = 70

# ════════════════════════════════════════════════════════════════════
# SHEET 5: Investment Analysis
# ════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Investment Analysis")
ws5.sheet_properties.tabColor = "8E44AD"

ws5.merge_cells("A1:F1")
ws5["A1"] = "AI Investment & Spend Analysis"
ws5["A1"].font = title_font

inv_headers = ["Firm", "Est. Annual AI\nSpend ($M)", "Total Tech\nBudget ($M)", "AI as % of\nTech Budget", "AI Spend as\n% of Revenue", "YoY AI Spend\nGrowth (est.)"]
for c, h in enumerate(inv_headers, 1):
    ws5.cell(row=3, column=c, value=h)
style_header_row(ws5, 3, len(inv_headers))

inv_data = [
    ["Morgan Stanley", 850, 5000, "17%", "6.5%", "35-40%"],
    ["BofA / Merrill Lynch", 900, 13000, "7%", "6.9%", "40-50%"],
    ["JP Morgan", 2000, 18000, "11%", "7.2%", "30-40%"],
    ["Goldman Sachs", 1200, 6000, "20%", "7.0%", "45-55%"],
    ["UBS", 800, 5500, "15%", "5.8%", "30-35%"],
    ["Charles Schwab", 500, 4000, "13%", "4.2%", "25-30%"],
    ["Fidelity", 600, 4500, "13%", "4.5%", "25-35%"],
    ["Vanguard", 400, 3500, "11%", "3.0%", "20-25%"],
    ["Wells Fargo", 700, 9000, "8%", "5.0%", "30-35%"],
    ["Raymond James", 120, 800, "15%", "2.5%", "20-25%"],
    ["Crestview Finance Group", 100, 700, "14%", "2.0%", "15-20%"],
    ["LPL Financial", 150, 900, "17%", "3.5%", "25-30%"],
    ["Wealthfront", 30, 80, "38%", "15.0%", "30-40%"],
    ["Betterment", 25, 65, "38%", "14.0%", "25-35%"],
    ["Robinhood", 200, 700, "29%", "12.0%", "40-50%"],
]

for i, row in enumerate(inv_data):
    r = 4 + i
    for c, val in enumerate(row, 1):
        ws5.cell(row=r, column=c, value=val)
        if c in (2, 3):
            ws5.cell(row=r, column=c).number_format = '#,##0'
            ws5.cell(row=r, column=c).font = data_font_blue
style_data_area(ws5, 4, 18, len(inv_headers))

# Totals row
ws5.cell(row=20, column=1, value="TOTAL / AVERAGE").font = bold_font
ws5.cell(row=20, column=2, value='=SUM(B4:B18)').font = Font(name="Arial", bold=True, size=11, color=MED_BLUE)
ws5.cell(row=20, column=2).number_format = '#,##0'
ws5.cell(row=20, column=3, value='=SUM(C4:C18)').font = Font(name="Arial", bold=True, size=11, color=MED_BLUE)
ws5.cell(row=20, column=3).number_format = '#,##0'
for c in range(1, len(inv_headers) + 1):
    ws5.cell(row=20, column=c).border = thin_border
    ws5.cell(row=20, column=c).fill = blue_fill

# Industry context
ws5.cell(row=22, column=1, value="INDUSTRY CONTEXT").font = subtitle_font
context_data = [
    ["Global wealth management IT spend (2024)", "$56.1B", "Source: Celent 2024"],
    ["AI/GenAI share of IT budget (2024)", "~16%", "Source: InvestmentNews/Wipro 2024 survey"],
    ["Projected AI share of IT budget (2028+)", "~32%+", "Expected to double in 3-5 years"],
    ["Global corporate AI investment (2024)", "$252.3B", "Source: Stanford HAI 2025 AI Index"],
    ["Global GenAI private investment (2024)", "$33.9B", "Source: Stanford HAI 2025 AI Index"],
    ["Cumulative AI investment since 2013", "$1.6T", "Source: Reuters"],
    ["AI wealth management market (2025 est.)", "$50B", "Source: AI Market Pulse"],
    ["% firms making moderate-large GenAI investment", "~75%", "Up from 40% in 2023; Source: Wipro survey"],
]
for i, (metric, val, source) in enumerate(context_data):
    r = 23 + i
    ws5.cell(row=r, column=1, value=metric).font = bold_font
    ws5.cell(row=r, column=1).border = thin_border
    ws5.cell(row=r, column=2, value=val).font = Font(name="Arial", bold=True, size=10, color=MED_BLUE)
    ws5.cell(row=r, column=2).alignment = center
    ws5.cell(row=r, column=2).border = thin_border
    ws5.cell(row=r, column=3, value=source).font = Font(name="Arial", italic=True, size=9, color="666666")
    ws5.cell(row=r, column=3).border = thin_border

for col in ["A", "B", "C", "D", "E", "F"]:
    ws5.column_dimensions[col].width = 22

# ════════════════════════════════════════════════════════════════════
# SHEET 6: Headcount & Focus
# ════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("Headcount & Focus")
ws6.sheet_properties.tabColor = "E74C3C"

ws6.merge_cells("A1:F1")
ws6["A1"] = "AI Headcount & Strategic Focus"
ws6["A1"].font = title_font

hc_headers = ["Firm", "Est. AI/ML\nHeadcount", "AI Focus\n(Client/Back/Balanced)", "Client-Facing\nAI", "Robo-\nAdvisory", "Key AI Roles & Notes"]
for c, h in enumerate(hc_headers, 1):
    ws6.cell(row=3, column=c, value=h)
style_header_row(ws6, 3, len(hc_headers))

hc_data = [
    ["Morgan Stanley", 1200, "Balanced", "Yes", "Yes", "OpenAI partnership team; 15K advisors using AI tools; dedicated AI engineering org"],
    ["BofA / Merrill Lynch", 1500, "Balanced", "Yes", "Yes", "1,200+ AI/ML patents; dedicated Erica team; 270+ models need ongoing ML engineers"],
    ["JP Morgan", 2000, "Balanced", "Yes", "No", "2,000+ AI/data experts — more than next 7 US banks combined; massive AI org under CTO"],
    ["Goldman Sachs", 500, "Back-Office", "Yes", "No", "~500 engineers/data scientists on AI; multi-model platform team; GitHub Copilot rollout"],
    ["UBS", 800, "Balanced", "Yes", "No", "AI Hub team; Azure partnership engineering; 30K employees using AI tools"],
    ["Charles Schwab", 600, "Balanced", "Yes", "Yes", "Knowledge Assistant team; Intelligent Portfolios ML team; NLP research group"],
    ["Fidelity", 700, "Balanced", "Yes", "Yes", "FCAT lab staff; 75+ teams involved in AI; actively hiring Principal AI/ML engineers (AWS)"],
    ["Vanguard", 500, "Client-Facing", "Yes", "Yes", "Quant fund ML team; Digital Advisor engineering; internal data science group"],
    ["Wells Fargo", 800, "Balanced", "Yes", "Yes", "Fargo platform team; Google Agentspace integration team; Pega implementation group"],
    ["Raymond James", 150, "Back-Office", "No", "No", "Small but growing AI team; focus on advisor tool enhancements; hiring in analytics"],
    ["Crestview Finance Group", 120, "Back-Office", "No", "No", "Digital transformation team; CRM AI integration; early-stage AI talent acquisition"],
    ["LPL Financial", 200, "Balanced", "Yes", "No", "Analytics platform team; advisor-facing AI tools; growing data science capability"],
    ["Wealthfront", 80, "Client-Facing", "Yes", "Yes", "Engineering-first culture; ML/RL for portfolio optimization; small but highly skilled team"],
    ["Betterment", 70, "Client-Facing", "Yes", "Yes", "ML optimization team; behavioral analytics; explainable AI for compliance engineering"],
    ["Robinhood", 250, "Client-Facing", "Yes", "Yes", "AI-first product team; Pluto acquisition added AI planning talent; massive data engineering"],
]

for i, row in enumerate(hc_data):
    r = 4 + i
    for c, val in enumerate(row, 1):
        ws6.cell(row=r, column=c, value=val)
style_data_area(ws6, 4, 18, len(hc_headers))

ws6.column_dimensions["A"].width = 24
ws6.column_dimensions["B"].width = 14
ws6.column_dimensions["C"].width = 22
ws6.column_dimensions["D"].width = 14
ws6.column_dimensions["E"].width = 12
ws6.column_dimensions["F"].width = 65

# ════════════════════════════════════════════════════════════════════
# SHEET 7: Sources & Methodology
# ════════════════════════════════════════════════════════════════════
ws7 = wb.create_sheet("Sources & Methodology")
ws7.sheet_properties.tabColor = "95A5A6"

ws7.merge_cells("A1:C1")
ws7["A1"] = "Research Sources & Methodology"
ws7["A1"].font = title_font

ws7.cell(row=3, column=1, value="METHODOLOGY").font = subtitle_font
methodology = [
    "This report synthesizes publicly available data from company press releases, earnings calls, SEC filings, industry analyst reports, and technology journalism.",
    "AI spend estimates are derived from: reported total tech budgets, industry benchmarks (AI = 8-16% of IT budget for traditional firms, 25-40% for digital-native), and specific company disclosures.",
    "AI headcount estimates use: LinkedIn job posting analysis, company disclosures, analyst estimates, and industry benchmarks (AI staff = 1-3% of total workforce for large firms).",
    "AI maturity scores are assigned based on: number of production AI use cases, breadth of deployment, investment levels, vendor partnerships, and public evidence of AI integration.",
    "All financial figures are estimates based on available data as of July 2025. Actual figures may differ. This is competitive intelligence research, not audited financial reporting.",
    "AUM figures are approximate and sourced from most recent public filings or industry rankings (2024-2025).",
]
for i, text in enumerate(methodology):
    ws7.cell(row=4 + i, column=1, value=text).font = data_font
    ws7.cell(row=4 + i, column=1).alignment = left_wrap
ws7.row_dimensions[4].height = 25

ws7.cell(row=12, column=1, value="KEY SOURCES").font = subtitle_font
sources = [
    ["Tearsheet", "The AI Reality Check — Q1 2025 Edition", "tearsheet.co"],
    ["Stanford HAI", "2025 AI Index Report — Economy Chapter", "hai.stanford.edu"],
    ["Oliver Wyman", "Morgan Stanley AI in Wealth & Asset Management", "oliverwyman.com"],
    ["OpenAI", "Morgan Stanley Case Study (AI Evals)", "openai.com"],
    ["CNBC", "Morgan Stanley OpenAI Debrief Launch; JPMorgan AI-powered megabank", "cnbc.com"],
    ["Celent", "Wealth Management IT Spending Forecasts 2023-2028", "celent.com"],
    ["InvestmentNews / Wipro", "Gen AI Gathers Momentum as Wealth Firms Scale", "investmentnews.com"],
    ["VentureBeat", "Wells Fargo AI Assistant — 245M Interactions", "venturebeat.com"],
    ["Google Cloud", "Wells Fargo Agentic AI with Agentspace", "cloud.google.com"],
    ["Microsoft", "UBS and Microsoft — Co-creating Banking with Azure AI", "microsoft.com"],
    ["Bank of America Newsroom", "AI Adoption by BofA Global Workforce", "newsroom.bankofamerica.com"],
    ["Financial Planning", "How Wealth Managers Are Responding to AI", "financial-planning.com"],
    ["Forbes Tech Council", "GenAI Revolutionizing Wealth Management (May 2025)", "forbes.com"],
    ["AI Expert Network", "Case Studies: JPMorgan, Goldman Sachs, Morgan Stanley, Wells Fargo", "aiexpert.network"],
    ["Reuters", "AI Investment Wave Analysis", "reuters.com"],
    ["J.D. Power", "2025 U.S. Financial Advisor Satisfaction Study", "jdpower.com"],
]

src_headers = ["Organization", "Publication / Report", "Domain"]
for c, h in enumerate(src_headers, 1):
    ws7.cell(row=13, column=c, value=h)
style_header_row(ws7, 13, 3)

for i, row in enumerate(sources):
    r = 14 + i
    for c, val in enumerate(row, 1):
        ws7.cell(row=r, column=c, value=val)
style_data_area(ws7, 14, 14 + len(sources) - 1, 3)

ws7.column_dimensions["A"].width = 28
ws7.column_dimensions["B"].width = 55
ws7.column_dimensions["C"].width = 30

# ── Save ──
output_path = "wealth-mgmt-ai-adoption-benchmarks-2024-2026.xlsx"
wb.save(output_path)
print(f"Saved: {output_path}")
