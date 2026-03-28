"""
AI in Financial Advisory & Wealth Management — Research Spreadsheet
Generates a multi-sheet Excel workbook with structured research data.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
import os

wb = Workbook()

# Color palette
NAVY = "0F2B46"
DARK_BLUE = "1B4F72"
STEEL = "2C3E50"
WHITE = "FFFFFF"
LIGHT_GRAY = "F2F4F6"
LIGHT_BLUE = "D6EAF8"
ACCENT_GREEN = "27AE60"
ACCENT_GOLD = "F39C12"
ACCENT_RED = "E74C3C"
MED_GRAY = "95A5A6"

# Reusable styles
hdr_font = Font(name="Arial", bold=True, color=WHITE, size=10)
hdr_fill = PatternFill("solid", fgColor=NAVY)
hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
body_font = Font(name="Arial", size=9)
body_align = Alignment(vertical="center", wrap_text=True)
num_align = Alignment(horizontal="right", vertical="center")
thin_border = Border(
    left=Side(style="thin", color="D5D8DC"),
    right=Side(style="thin", color="D5D8DC"),
    top=Side(style="thin", color="D5D8DC"),
    bottom=Side(style="thin", color="D5D8DC"),
)
alt_fill = PatternFill("solid", fgColor=LIGHT_GRAY)
blue_fill = PatternFill("solid", fgColor=LIGHT_BLUE)

title_font = Font(name="Georgia", bold=True, size=16, color=NAVY)
subtitle_font = Font(name="Arial", size=10, color=MED_GRAY)
section_font = Font(name="Arial", bold=True, size=12, color=DARK_BLUE)
kpi_label_font = Font(name="Arial", size=9, color=MED_GRAY)
kpi_value_font = Font(name="Georgia", bold=True, size=20, color=NAVY)

def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = thin_border

def style_row(ws, row, ncols, alt=False):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = body_font
        cell.alignment = body_align
        cell.border = thin_border
        if alt:
            cell.fill = alt_fill

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def write_title(ws, title, subtitle=None):
    ws.merge_cells("A1:J1")
    c = ws.cell(row=1, column=1, value=title)
    c.font = title_font
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 38
    if subtitle:
        ws.merge_cells("A2:J2")
        c2 = ws.cell(row=2, column=1, value=subtitle)
        c2.font = subtitle_font
        c2.alignment = Alignment(horizontal="left")

# ============================================================
# SHEET 1 — Market Overview & Key Metrics
# ============================================================
ws1 = wb.active
ws1.title = "Market Overview"
write_title(ws1, "AI in Wealth Management — Market Overview",
            "Research compiled from PwC, McKinsey, Deloitte, and industry reports (2024–2025)")

# KPI Cards
kpi_data = [
    ("AI in WM Market Size (2024)", "$3.4–3.7B"),
    ("Projected Market (2030)", "$8.9–13.5B"),
    ("CAGR (2024–2030)", "10.8–22%"),
    ("Robo-Advisory AUM (2024)", "$1.2T"),
    ("Firms Using/Planning AI", "91%+"),
    ("Expected Revenue Lift", "12% by 2028"),
]

for i, (label, value) in enumerate(kpi_data):
    col = i + 1
    ws1.column_dimensions[get_column_letter(col)].width = 22
    ws1.cell(row=4, column=col, value=label).font = kpi_label_font
    ws1.cell(row=4, column=col).alignment = Alignment(horizontal="center")
    ws1.cell(row=4, column=col).border = thin_border
    vc = ws1.cell(row=5, column=col, value=value)
    vc.font = kpi_value_font
    vc.alignment = Alignment(horizontal="center")
    vc.border = thin_border
ws1.row_dimensions[5].height = 40

# Market Size Table
r = 7
ws1.cell(row=r, column=1, value="Market Size Estimates by Segment").font = section_font
r = 8
mkt_hdrs = ["Segment", "2024 Est.", "2025 Est.", "2030 Proj.", "CAGR", "Source"]
for i, h in enumerate(mkt_hdrs, 1):
    ws1.cell(row=r, column=i, value=h)
set_col_widths(ws1, [32, 16, 16, 16, 12, 34])
style_header(ws1, r, len(mkt_hdrs))

mkt_data = [
    ["AI in Wealth Management (core)", "$3.4–3.7B", "$3.4–4.2B", "$8.9–13.5B", "10.8–22%", "ConsaInsights / Dataintelo / StrategyMRC"],
    ["AI-Powered WM Solutions", "$1.6B", "$1.8B", "$8.9B", "20%+", "Future Market Insights"],
    ["Wealth Tech Solutions (incl. robo)", "$6.5B", "$6.9B", "$13.5B", "10.8%", "Mordor Intelligence"],
    ["Robo-Advisory (AUM)", "$1.09T", "$1.2T", "$2.5T (est.)", "~15%", "Condor Capital / AAII"],
    ["Robo-Advisory (Revenue)", "$55B", "$62B", "$120B (est.)", "~12%", "Industry estimates"],
    ["Global WM Market (total AUM)", "$128T", "$135T", "$2.95T rev.", "5–7%", "PwC AWM Report 2024"],
    ["AI Application Layer (finance)", "—", "$250B", "$395B", "~18%", "UBS GWM Research"],
]

for i, row in enumerate(mkt_data):
    rr = r + 1 + i
    for c, val in enumerate(row, 1):
        ws1.cell(row=rr, column=c, value=val)
    style_row(ws1, rr, len(mkt_hdrs), alt=(i % 2 == 1))

# Adoption Rates Table
r2 = r + len(mkt_data) + 3
ws1.cell(row=r2, column=1, value="AI Adoption Rates in Wealth Management").font = section_font
r2 += 1
adopt_hdrs = ["Metric", "Value", "Year", "Source"]
for i, h in enumerate(adopt_hdrs, 1):
    ws1.cell(row=r2, column=i, value=h)
style_header(ws1, r2, len(adopt_hdrs))

adopt_data = [
    ["Asset managers using/planning AI", "91%", "2024", "PwC AWM Report"],
    ["Firms expect AI-driven revenue growth", "80%", "2024", "PwC AWM Report"],
    ["AI seen as most transformational tech", "73%", "2024", "Fidelity AI Pulse Survey"],
    ["WM firms piloting/scaling GenAI", "67%+", "2025", "Fidelity AI Pulse Survey"],
    ["Firms plan to boost AI spend ≥10%", "91%", "2024–2027", "Fidelity / Morningstar"],
    ["Morgan Stanley advisor AI adoption", "98%", "2024", "InvestmentNews"],
    ["Schwab Knowledge Assistant adoption", "90%", "2024", "AboutSchwab / Emerj"],
    ["Retail investor AI advisory penetration", "80% (projected)", "2028", "Industry estimates"],
    ["Robo-advisory share of WealthTech", "33.6%", "2024", "Mordor Intelligence"],
    ["Banks' share of WealthTech market", "42.2%", "2024", "Mordor Intelligence"],
]

for i, row in enumerate(adopt_data):
    rr = r2 + 1 + i
    for c, val in enumerate(row, 1):
        ws1.cell(row=rr, column=c, value=val)
    style_row(ws1, rr, len(adopt_hdrs), alt=(i % 2 == 1))

ws1.freeze_panes = "A4"

# ============================================================
# SHEET 2 — Firm-by-Firm AI Initiatives
# ============================================================
ws2 = wb.create_sheet("Firm AI Initiatives")
write_title(ws2, "Firm-by-Firm AI Initiatives in Wealth Management",
            "Detailed breakdown of AI deployments at top 10 wealth management firms")

hdrs2 = [
    "Firm", "Est. Total AUM", "Robo-Advisory AUM", "Key AI Initiatives",
    "Client-Facing AI?", "Robo-Advisory?", "GenAI Tools Deployed",
    "Est. Annual AI Spend", "Est. AI Headcount", "AI Partner(s)", "Notable Metric"
]
widths2 = [20, 16, 18, 50, 14, 14, 36, 18, 16, 20, 40]
set_col_widths(ws2, widths2)

r = 4
for i, h in enumerate(hdrs2, 1):
    ws2.cell(row=r, column=i, value=h)
style_header(ws2, r, len(hdrs2))

firms = [
    [
        "Morgan Stanley", "~$4.6T", "N/A (hybrid only)",
        "AI @ MS Assistant (GPT-4 knowledge base); AI @ MS Debrief (meeting notes/follow-up); AskResearchGPT; Next Best Action engine; Predictive client analytics",
        "Yes (via advisor)", "No (hybrid PAS)", "AI Assistant, Debrief, AskResearchGPT",
        "$1.5B+ (tech total)", "~2,000+", "OpenAI (exclusive WM partner)",
        "98% advisor adoption; 30 min saved/meeting; doc retrieval 20%→80%"
    ],
    [
        "Merrill Lynch / BofA", "~$3.5T", "Merrill Guided Investing",
        "Erica virtual assistant (3B+ interactions); Ask Merrill (advisor tool); AI tax harvesting; Predictive client insights; CashPro AI for treasury; 270+ AI models in production",
        "Yes (Erica)", "Yes (Merrill GI)", "Erica, Ask Merrill, CashPro AI",
        "$4B/yr (BofA total AI)", "~3,500+", "In-house + vendors",
        "50M Erica users; 58M monthly interactions; 98% auto-resolution; 42% chat reduction"
    ],
    [
        "Charles Schwab", "~$8.5T (client assets)", "$89.5B",
        "Schwab Intelligent Portfolios (robo); Schwab Knowledge Assistant (GenAI for reps); NLP Investing Themes (40+ themes); ML fraud detection; Schwab Research Assistant",
        "Yes (SIP)", "Yes (SIP, SIP Premium)", "Knowledge Assistant, Investing Themes",
        "$1.5B+ (tech total)", "~1,500+", "In-house proprietary",
        "90% employee adoption of Knowledge Assistant; 25% cost/account reduction over decade"
    ],
    [
        "Fidelity Investments", "~$4.5T", "Fidelity Go (est. $30B+)",
        "Fidelity Go robo-advisor; AI-powered customer service NLP; Advisor productivity tools (note-taking, meeting prep); Predictive analytics for portfolio mgrs",
        "Yes (Fidelity Go)", "Yes (Fidelity Go)", "Customer service AI, advisor tools",
        "$2.5B+ (tech total)", "~2,000+", "In-house + partnerships",
        "0.35% fee (>$25K); no min balance; Flex funds (0% ER); 2/3 of WM firms piloting GenAI per Fidelity survey"
    ],
    [
        "Vanguard", "~$8.6T", "$365.1B (PAS + Digital)",
        "Vanguard Digital Advisor (robo); Personal Advisor Services (hybrid AI+human); Automated rebalancing; Goal-based planning algorithms; Risk profiling ML",
        "Yes (Digital Advisor)", "Yes (VDA, PAS)", "Digital Advisor platform",
        "$1B+ (est. tech)", "~1,000+", "In-house proprietary",
        "Largest robo AUM ($365B); 0.15% digital fee; $3K min; industry-leading low-cost model"
    ],
    [
        "Wealthfront", "~$50B", "$35.3B",
        "Fully automated robo-advisor; 'Path' financial planning AI; Continuous tax-loss harvesting; Direct indexing; Risk parity portfolios; Cash flow forecasting",
        "Yes (fully digital)", "Yes (core product)", "Path planning engine, TLH automation",
        "$50–100M (est.)", "~200–300", "In-house proprietary",
        "$35.3B AUM; 0.25% fee; $500 min; pioneer in automated TLH and direct indexing"
    ],
    [
        "Betterment", "~$56B", "$56.4B",
        "Automated portfolio management; Tax-loss harvesting; Goal-based planning AI; Tax-Coordinated Portfolio; RetireGuide; Charitable giving optimization",
        "Yes (fully digital)", "Yes (core product)", "RetireGuide, TLH engine, goal planner",
        "$50–100M (est.)", "~300–400", "In-house + Goldman (B2B)",
        "$56.4B AUM; 0.25% digital / 0.65% premium; absorbed Marcus Invest & United Income clients"
    ],
    [
        "Goldman Sachs", "~$2.8T (AWM)", "Marcus (wound down → Betterment)",
        "$4B AI fintech investment; Marcus AI platform (portfolio adjustment, risk forecasting); GS AI Assistant (GenAI coding/productivity); IndexGPT (thematic baskets); AI-driven advisory",
        "Yes (Marcus AI)", "No (exited retail robo)", "Marcus AI, GS AI Assistant, IndexGPT",
        "$4B (announced 2025)", "~2,000+", "In-house proprietary",
        "18% Q1 2025 advisory revenue increase; 40% more client data processing vs peers; targeting 20% AI WM market share"
    ],
    [
        "JP Morgan", "~$3.9T (AWM)", "N/A (no standalone robo)",
        "LLM Suite (200K+ daily users); Coach AI & Smart Monitor (advisor tools); IndexGPT (thematic investing); EVEE Q&A agent; 450+ AI use cases; Agentic AI workflows",
        "Yes (IndexGPT)", "No", "LLM Suite, Coach AI, Smart Monitor, IndexGPT, EVEE",
        "$18B/yr (total tech; ~$3B+ AI)", "~3,000+ AI/ML", "OpenAI, Anthropic, in-house",
        "20% gross sales increase from AI; 3.4x advisor productivity; 83% research time reduction; 50% more clients/advisor"
    ],
    [
        "UBS", "~$5.7T (invested assets)", "N/A (no standalone robo)",
        "UBS Red (AI advisor assistant on Azure); My Way (customizable digital WM); 55K Microsoft Copilot licenses; Data mesh + RAG for unstructured data; AI Hub governance framework",
        "Yes (My Way)", "No (digital hybrid only)", "UBS Red, My Way, MS Copilot",
        "$2B+ (est. tech total)", "~1,500+", "Microsoft (Azure OpenAI)",
        "60K+ digitized investment docs; multi-language support; integrated Credit Suisse client migration with AI"
    ],
]

for i, row in enumerate(firms):
    rr = r + 1 + i
    for c, val in enumerate(row, 1):
        ws2.cell(row=rr, column=c, value=val)
    style_row(ws2, rr, len(hdrs2), alt=(i % 2 == 1))

ws2.freeze_panes = "B5"
ws2.auto_filter.ref = f"A4:{get_column_letter(len(hdrs2))}{r + len(firms)}"

# ============================================================
# SHEET 3 — Robo-Advisory Comparison
# ============================================================
ws3 = wb.create_sheet("Robo-Advisory Comparison")
write_title(ws3, "Robo-Advisory Platform Comparison",
            "Fee structures, AUM, capabilities, and minimums for leading robo-advisors (2024)")

hdrs3 = [
    "Platform", "Parent Firm", "AUM (2024)", "Digital Fee", "Premium/Hybrid Fee",
    "Account Min.", "Tax-Loss Harvesting?", "Human Advisor Access?",
    "Key Differentiator", "Fund Types"
]
widths3 = [24, 18, 16, 14, 20, 14, 18, 18, 36, 24]
set_col_widths(ws3, widths3)

r = 4
for i, h in enumerate(hdrs3, 1):
    ws3.cell(row=r, column=i, value=h)
style_header(ws3, r, len(hdrs3))

robos = [
    ["Vanguard Digital Advisor", "Vanguard", "$365.1B (combined)", "0.15%", "0.30% (PAS)", "$3,000",
     "Yes", "Yes (PAS tier)", "Largest AUM; lowest hybrid fee", "Vanguard ETFs & mutual funds"],
    ["Schwab Intelligent Portfolios", "Charles Schwab", "$89.5B", "0% (no advisory fee)", "$30/mo + $300 setup",
     "$5,000", "Yes (>$50K)", "Yes (Premium)", "Zero advisory fee model; cash drag trade-off",
     "Schwab-selected ETFs"],
    ["Betterment Digital", "Betterment", "$56.4B", "0.25%", "0.65% (Premium, $100K min)",
     "$0", "Yes", "Yes (Premium tier)", "Goal-based planning; absorbed Marcus/United Income clients",
     "Diversified ETF portfolio"],
    ["Wealthfront", "Wealthfront Inc.", "$35.3B", "0.25%", "N/A (digital only)",
     "$500", "Yes", "No", "Path planning tool; direct indexing pioneer", "ETFs, direct indexing"],
    ["Fidelity Go", "Fidelity Investments", "~$30B+ (est.)", "0% (<$25K) / 0.35%", "Coaching (>$25K)",
     "$0 ($10 to invest)", "No", "Yes (>$25K)", "Zero-ER Flex funds; no min balance",
     "Fidelity Flex mutual funds (0% ER)"],
    ["SoFi Automated Investing", "SoFi", "~$5B (est.)", "0%", "N/A",
     "$1", "No", "Yes (SoFi members)", "Free robo with SoFi ecosystem", "ETFs"],
    ["Merrill Guided Investing", "Bank of America", "~$30B (est.)", "0.45%", "0.85% (with advisor)",
     "$1,000", "No", "Yes (w/ advisor tier)", "BofA/Merrill ecosystem integration; Erica access",
     "ETFs, mutual funds"],
]

for i, row in enumerate(robos):
    rr = r + 1 + i
    for c, val in enumerate(row, 1):
        ws3.cell(row=rr, column=c, value=val)
    style_row(ws3, rr, len(hdrs3), alt=(i % 2 == 1))

ws3.freeze_panes = "B5"

# ============================================================
# SHEET 4 — AI Use Cases by Category
# ============================================================
ws4 = wb.create_sheet("AI Use Cases")
write_title(ws4, "AI Use Cases in Wealth Management",
            "Categorized breakdown of AI applications across the industry")

hdrs4 = [
    "Use Case Category", "Specific Application", "Technology", "Firms Deploying",
    "Maturity", "Client Impact", "Est. % of AI Spend"
]
widths4 = [24, 36, 24, 36, 14, 36, 16]
set_col_widths(ws4, widths4)

r = 4
for i, h in enumerate(hdrs4, 1):
    ws4.cell(row=r, column=i, value=h)
style_header(ws4, r, len(hdrs4))

use_cases = [
    ["Robo-Advisory", "Automated portfolio construction & rebalancing",
     "ML, optimization algos", "Vanguard, Schwab, Betterment, Wealthfront, Fidelity",
     "Mature", "Democratized investing; lower fees (0–0.35%)", "15%"],
    ["Robo-Advisory", "Goal-based financial planning AI",
     "ML, Monte Carlo simulation", "Betterment, Wealthfront (Path), Fidelity Go",
     "Mature", "Automated retirement/college planning", "5%"],
    ["Tax Optimization", "Continuous tax-loss harvesting",
     "ML, real-time monitoring", "Wealthfront, Betterment, Schwab, Vanguard",
     "Mature", "After-tax return improvement 0.5–1.5%/yr", "8%"],
    ["Tax Optimization", "Direct indexing with AI lot selection",
     "ML optimization", "Wealthfront, Parametric (MS), Aperio (BlackRock)",
     "Growing", "Personalized tax alpha for HNW clients", "4%"],
    ["Portfolio Management", "AI-assisted factor-based rebalancing",
     "ML, factor models", "JPMorgan, Goldman Sachs, Morgan Stanley",
     "Growing", "Better risk-adjusted returns", "7%"],
    ["Portfolio Management", "Risk optimization & scenario analysis",
     "ML, stress testing", "JPMorgan (Smart Monitor), Goldman Sachs, UBS",
     "Growing", "83% research time reduction (JPM)", "5%"],
    ["Client Communications", "AI meeting notes & follow-ups",
     "GenAI, NLP, speech-to-text", "Morgan Stanley (Debrief)",
     "Emerging", "30 min saved per meeting; higher quality notes", "6%"],
    ["Client Communications", "AI-generated client reports & emails",
     "GenAI (GPT-4, Claude)", "Morgan Stanley, JPMorgan, UBS",
     "Emerging", "Faster, more personalized client outreach", "5%"],
    ["Client Communications", "Virtual assistant / chatbot",
     "NLP, GenAI", "BofA (Erica), Schwab (Knowledge Asst), JPM (EVEE)",
     "Mature", "98% auto-resolution (Erica); 50M users", "8%"],
    ["Sentiment Analysis", "Client communication sentiment tracking",
     "NLP, sentiment models", "Morgan Stanley (Next Best Action), JPMorgan",
     "Growing", "Proactive churn prevention; relationship deepening", "4%"],
    ["Sentiment Analysis", "Market sentiment from news/social media",
     "NLP, alternative data", "Goldman Sachs, JPMorgan, BlackRock",
     "Growing", "Alpha generation from alternative data signals", "3%"],
    ["Financial Planning", "ML-enhanced Monte Carlo simulations",
     "ML, stochastic models", "Fidelity, Schwab, Vanguard, Betterment",
     "Growing", "More accurate retirement projections", "4%"],
    ["Financial Planning", "Cash flow forecasting AI",
     "ML, time series", "Wealthfront (Path), BofA (CashPro), Fidelity",
     "Growing", "Better liquidity planning for clients", "3%"],
    ["Predictive Analytics", "Client churn prediction",
     "ML classification models", "Morgan Stanley, JPMorgan, UBS, BofA",
     "Growing", "Proactive retention; reduced attrition 15–25%", "5%"],
    ["Predictive Analytics", "Next-best-action engines",
     "ML, reinforcement learning", "Morgan Stanley, JPMorgan, Goldman Sachs",
     "Growing", "20% gross sales increase (JPM); personalized engagement", "5%"],
    ["Predictive Analytics", "Client lifetime value prediction",
     "ML regression models", "JPMorgan, Morgan Stanley, UBS",
     "Emerging", "Better resource allocation to high-value clients", "3%"],
    ["Compliance & Risk", "AI-powered AML/KYC",
     "ML, graph analytics", "All major firms (JPM COIN platform)",
     "Mature", "Billions in fraud prevented annually", "6%"],
    ["Compliance & Risk", "Regulatory document review",
     "NLP, GenAI", "JPMorgan, Goldman Sachs, Morgan Stanley",
     "Growing", "40% cost reduction in onboarding (JPM)", "4%"],
]

for i, row in enumerate(use_cases):
    rr = r + 1 + i
    for c, val in enumerate(row, 1):
        ws4.cell(row=rr, column=c, value=val)
    style_row(ws4, rr, len(hdrs4), alt=(i % 2 == 1))

# Conditional formatting on Maturity
ws4.conditional_formatting.add(
    f"E5:E{r + len(use_cases)}",
    CellIsRule(operator="equal", formula=['"Mature"'],
              fill=PatternFill("solid", fgColor="D5F5E3"),
              font=Font(color=ACCENT_GREEN, bold=True))
)
ws4.conditional_formatting.add(
    f"E5:E{r + len(use_cases)}",
    CellIsRule(operator="equal", formula=['"Growing"'],
              fill=PatternFill("solid", fgColor="FEF9E7"),
              font=Font(color=ACCENT_GOLD, bold=True))
)
ws4.conditional_formatting.add(
    f"E5:E{r + len(use_cases)}",
    CellIsRule(operator="equal", formula=['"Emerging"'],
              fill=PatternFill("solid", fgColor="FADBD8"),
              font=Font(color=ACCENT_RED, bold=True))
)

ws4.freeze_panes = "B5"

# ============================================================
# SHEET 5 — AI Spend Distribution
# ============================================================
ws5 = wb.create_sheet("AI Spend Distribution")
write_title(ws5, "AI Use Case Distribution & Spend Allocation",
            "Estimated breakdown of AI investment across functional areas in wealth management")

hdrs5 = ["Functional Area", "Est. % of AI Spend", "Key Applications", "Primary Beneficiary", "Maturity Level"]
widths5 = [24, 18, 50, 20, 16]
set_col_widths(ws5, widths5)

r = 4
for i, h in enumerate(hdrs5, 1):
    ws5.cell(row=r, column=i, value=h)
style_header(ws5, r, len(hdrs5))

dist_data = [
    ["Advisory & Portfolio Mgmt", 0.35, "Robo-advisory, portfolio rebalancing, factor investing, direct indexing, goal planning",
     "Clients & Advisors", "Mature/Growing"],
    ["Client Service & Engagement", 0.22, "Virtual assistants (Erica), chatbots, NLP query handling, client onboarding AI",
     "Clients", "Mature"],
    ["Compliance & Risk", 0.18, "AML/KYC automation, regulatory doc review, fraud detection, trade surveillance",
     "Operations/Legal", "Mature"],
    ["Operations & Efficiency", 0.15, "Meeting notes (Debrief), email drafting, report generation, knowledge search, coding assistants",
     "Advisors & Staff", "Emerging/Growing"],
    ["Analytics & Insights", 0.10, "Client churn prediction, sentiment analysis, next-best-action, LTV modeling, market signals",
     "Advisors & Mgmt", "Growing"],
]

for i, row in enumerate(dist_data):
    rr = r + 1 + i
    for c, val in enumerate(row, 1):
        ws5.cell(row=rr, column=c, value=val)
    ws5.cell(row=rr, column=2).number_format = "0%"
    style_row(ws5, rr, len(hdrs5), alt=(i % 2 == 1))

# Total row
tr = r + len(dist_data) + 1
ws5.cell(row=tr, column=1, value="TOTAL").font = Font(name="Arial", bold=True, size=10)
ws5.cell(row=tr, column=2, value=f"=SUM(B5:B{tr-1})")
ws5.cell(row=tr, column=2).number_format = "0%"
ws5.cell(row=tr, column=2).font = Font(name="Arial", bold=True, size=10)
for c in range(1, len(hdrs5) + 1):
    ws5.cell(row=tr, column=c).border = Border(
        top=Side(style="double", color=NAVY), bottom=Side(style="double", color=NAVY))

# Per-firm spend estimates
r2 = tr + 3
ws5.cell(row=r2, column=1, value="Estimated Annual AI/Tech Spend by Firm").font = section_font
r2 += 1
spend_hdrs = ["Firm", "Total Tech Budget (est.)", "AI-Specific Spend (est.)", "AI as % of Tech", "Key AI Investment"]
for i, h in enumerate(spend_hdrs, 1):
    ws5.cell(row=r2, column=i, value=h)
style_header(ws5, r2, len(spend_hdrs))

spend_data = [
    ["JP Morgan", "$18B/yr", "~$3–4B", "17–22%", "$18B tech budget; 450+ AI use cases; LLM Suite"],
    ["Bank of America", "$12B/yr", "~$4B (AI/digital)", "33%", "$4B/yr AI commitment; 270+ models; Erica"],
    ["Goldman Sachs", "$5B/yr", "~$4B (fintech AI bet)", "80% (one-time)", "$4B announced AI fintech investment in 2025"],
    ["Morgan Stanley", "$4B/yr", "~$1.5B", "38%", "Exclusive OpenAI WM partnership; AI suite"],
    ["Charles Schwab", "$3B/yr", "~$1–1.5B", "33–50%", "Knowledge Assistant; SIP platform; NLP themes"],
    ["UBS", "$4B/yr", "~$2B+", "50%+", "55K Copilot licenses; Azure partnership; UBS Red"],
    ["Fidelity", "$4B/yr", "~$2–2.5B", "50–63%", "Fidelity Go; advisor tools; AI Pulse program"],
    ["Vanguard", "$2B/yr", "~$500M–1B", "25–50%", "Digital Advisor platform; PAS hybrid model"],
    ["Wealthfront", "$100–150M/yr", "~$50–100M", "50–67%", "Path engine; TLH automation; all-digital"],
    ["Betterment", "$100–150M/yr", "~$50–100M", "50–67%", "Goal planning AI; TLH; B2B platform"],
]

for i, row in enumerate(spend_data):
    rr = r2 + 1 + i
    for c, val in enumerate(row, 1):
        ws5.cell(row=rr, column=c, value=val)
    style_row(ws5, rr, len(spend_hdrs), alt=(i % 2 == 1))

# ============================================================
# SHEET 6 — GenAI Tools Detail
# ============================================================
ws6 = wb.create_sheet("GenAI Tools Detail")
write_title(ws6, "Generative AI Tools Deployed in Wealth Management",
            "Specific GenAI products shipped or announced by major firms (2023–2025)")

hdrs6 = [
    "Firm", "Tool Name", "Launch Year", "Technology", "Use Case",
    "User Base", "Key Metric", "Status"
]
widths6 = [18, 28, 12, 20, 36, 24, 40, 12]
set_col_widths(ws6, widths6)

r = 4
for i, h in enumerate(hdrs6, 1):
    ws6.cell(row=r, column=i, value=h)
style_header(ws6, r, len(hdrs6))

tools_data = [
    ["Morgan Stanley", "AI @ MS Assistant", "2023", "OpenAI GPT-4",
     "Knowledge search, research synthesis, process guidance",
     "16,000+ advisors", "98% advisor adoption; doc retrieval 20%→80%", "Live"],
    ["Morgan Stanley", "AI @ MS Debrief", "2024", "OpenAI GPT-4",
     "Meeting notes, summaries, follow-up emails, CRM integration",
     "16,000+ advisors", "30 min saved/meeting; higher quality than human notes", "Live"],
    ["Morgan Stanley", "AskResearchGPT", "2024", "OpenAI GPT-4",
     "Synthesize institutional research for advisors",
     "Advisors & analysts", "Instant research synthesis across thousands of reports", "Live"],
    ["Bank of America", "Erica", "2018 (AI upgrades ongoing)", "NLP, ML, GenAI",
     "Virtual financial assistant for clients",
     "50M clients", "3B+ interactions; 58M/month; 98% auto-resolution", "Live"],
    ["Bank of America", "Ask Merrill", "2023", "NLP, GenAI",
     "Advisor-facing query tool for client portfolio mgmt",
     "Merrill advisors", "Covers 50+ investment topics; integrated with Erica", "Live"],
    ["Bank of America", "CashPro AI", "2023", "NLP, ML",
     "Treasury/business banking AI assistant",
     "Business clients", "42% reduction in live chat volume", "Live"],
    ["Charles Schwab", "Schwab Knowledge Assistant", "2024", "GenAI (proprietary)",
     "Internal tool for reps to answer complex client queries",
     "All Schwab reps", "90% adoption; faster handling times", "Live"],
    ["Charles Schwab", "Investing Themes", "2024", "NLP models",
     "Thematic investing via NLP analysis of unstructured data",
     "Retail investors", "40+ themed portfolios; no advisory fee", "Live"],
    ["JP Morgan", "LLM Suite", "2023–2024", "OpenAI, Anthropic, in-house",
     "Firm-wide GenAI platform for all business lines",
     "200,000+ employees", "Updated every 8 weeks; 30–40% productivity gain", "Live"],
    ["JP Morgan", "Coach AI", "2024", "ML, GenAI",
     "Real-time research and recommendations for advisors",
     "Wealth advisors", "3.4x advisor productivity increase", "Live"],
    ["JP Morgan", "Smart Monitor", "2024", "ML analytics",
     "Portfolio monitoring and research automation",
     "Portfolio managers", "83% research time reduction", "Live"],
    ["JP Morgan", "IndexGPT", "2024", "GPT-4, NLP",
     "Thematic investment basket creation for clients",
     "Clients & advisors", "Real-time market scanning for themes", "Live"],
    ["JP Morgan", "EVEE Q&A", "2024", "GenAI agent",
     "Intelligent Q&A for call centers and client service",
     "Service teams", "Instant specialist answers from policy/data", "Live"],
    ["Goldman Sachs", "Marcus AI", "2024–2025", "Proprietary LLM",
     "Personalized portfolio adjustments & risk forecasting",
     "Wealth clients", "40% more data processing vs peers; 18% rev increase", "Live"],
    ["Goldman Sachs", "GS AI Assistant", "2024", "GenAI",
     "Coding assistant and productivity tool for staff",
     "All employees", "Firm-wide deployment for operational efficiency", "Live"],
    ["UBS", "UBS Red", "2024", "Azure OpenAI",
     "AI advisor assistant with knowledge base search",
     "Client advisors", "60K+ digitized docs; multi-language; conceptual search", "Live"],
    ["UBS", "My Way", "2024", "AI/digital platform",
     "Customizable modular digital wealth management",
     "Clients", "Client-defined investment parameters with AI guidance", "Live"],
    ["Fidelity", "Fidelity Go", "2016 (AI enhanced)", "ML algorithms",
     "Robo-advisor with automated rebalancing",
     "Retail investors", "0% fee <$25K; coaching >$25K; Flex funds", "Live"],
    ["Vanguard", "Digital Advisor", "2020 (enhanced)", "ML algorithms",
     "Low-cost robo-advisory with goal planning",
     "Retail investors", "$365B AUM; 0.15% fee; $3K min", "Live"],
    ["Wealthfront", "Path", "2017 (ML enhanced)", "ML, Monte Carlo",
     "Financial planning engine with scenario modeling",
     "All clients", "Integrated with all accounts; free with 0.25% AUM", "Live"],
    ["Betterment", "RetireGuide", "2023", "ML planning",
     "Retirement planning with personalized projections",
     "All clients", "Goal-based with tax optimization integration", "Live"],
]

for i, row in enumerate(tools_data):
    rr = r + 1 + i
    for c, val in enumerate(row, 1):
        ws6.cell(row=rr, column=c, value=val)
    style_row(ws6, rr, len(hdrs6), alt=(i % 2 == 1))

ws6.freeze_panes = "B5"
ws6.auto_filter.ref = f"A4:{get_column_letter(len(hdrs6))}{r + len(tools_data)}"

# ============================================================
# SHEET 7 — Sources & Methodology
# ============================================================
ws7 = wb.create_sheet("Sources")
write_title(ws7, "Sources & Methodology", "Data sources and methodology notes")

set_col_widths(ws7, [8, 50, 60])

r = 4
ws7.cell(row=r, column=1, value="#").font = Font(name="Arial", bold=True, size=10)
ws7.cell(row=r, column=2, value="Source").font = Font(name="Arial", bold=True, size=10)
ws7.cell(row=r, column=3, value="Data Used").font = Font(name="Arial", bold=True, size=10)
style_header(ws7, r, 3)

sources = [
    [1, "PwC 2024 Asset & Wealth Management Report", "Adoption rates, revenue projections, industry trends"],
    [2, "Condor Capital Robo Report (Q2 2025)", "Robo-advisory AUM figures for all platforms"],
    [3, "AAII 2024 Robo-Advice Landscape", "Fee structures, platform comparisons, industry shifts"],
    [4, "Mordor Intelligence Wealth Tech Report", "Market sizing, segment breakdown, CAGR"],
    [5, "ConsaInsights AI in WM Market Report", "AI market sizing ($3.4–3.7B in 2024)"],
    [6, "Dataintelo Wealth Management AI Report", "Market projections to 2033"],
    [7, "Future Market Insights AI-Powered WM", "Solution market sizing and forecasts"],
    [8, "StrategyMRC AI in Wealth Management", "Broader market definition ($25.4B)"],
    [9, "Fidelity 2025 AI Pulse Survey", "Firm adoption rates, GenAI piloting statistics"],
    [10, "CNBC — Morgan Stanley AI reports (2024)", "AI @ MS Debrief launch, advisor adoption"],
    [11, "BusinessWire — Morgan Stanley (June 2024)", "Debrief announcement, efficiency metrics"],
    [12, "InvestmentNews — Morgan Stanley AI (2024)", "98% advisor adoption figure"],
    [13, "BofA Newsroom (2025)", "Erica 3B interactions, 50M users, CashPro AI"],
    [14, "The Financial Brand — BofA AI Spend", "$4B annual AI investment; 270+ models"],
    [15, "Emerj — Charles Schwab AI Use Cases", "Knowledge Assistant, thematic investing"],
    [16, "AboutSchwab — AI on the Line", "90% adoption; service cost reduction"],
    [17, "CNBC — JPMorgan AI Megabank (2025)", "$18B tech budget; AI-first transformation"],
    [18, "AI Expert Network — JPMorgan Case Study", "450 use cases; 20% sales increase"],
    [19, "AI Market Pulse — Goldman Sachs $4B AI", "Marcus AI; fintech investment details"],
    [20, "UBS / Microsoft Partnership (2025)", "UBS Red; Azure deployment; 55K Copilot licenses"],
    [21, "Morningstar — Fidelity AI in WM", "AI use cases, advisor tool deployment"],
    [22, "Virtue Market Research — WM Market", "Global WM market to $2.95T by 2030"],
    [23, "UBS GWM AI Investment Framework", "$395B AI application market by 2027"],
]

for i, row in enumerate(sources):
    rr = r + 1 + i
    for c, val in enumerate(row, 1):
        ws7.cell(row=rr, column=c, value=val)
    style_row(ws7, rr, 3, alt=(i % 2 == 1))

# Methodology notes
rn = r + len(sources) + 2
ws7.cell(row=rn, column=1, value="Methodology Notes").font = section_font
notes = [
    "• AUM figures are approximate and sourced from public filings, press releases, and industry reports (2024 year-end where available).",
    "• AI spend estimates are derived from public disclosures, analyst reports, and industry benchmarks. Actual figures may vary.",
    "• AI headcount estimates are based on LinkedIn data, job postings, and public statements. Includes AI/ML engineers, data scientists, and AI product roles.",
    "• Use case distribution percentages are estimated based on industry surveys (PwC, Deloitte, McKinsey) and public firm disclosures.",
    "• Market size ranges reflect different analyst definitions and scope (narrow AI-specific vs. broader wealth tech).",
    "• 'Est.' and '~' prefix indicates estimated or approximate values based on available data.",
    "• Research date: October 2025. Data reflects information available through this date.",
]
for i, note in enumerate(notes):
    ws7.cell(row=rn + 1 + i, column=1, value=note).font = Font(name="Arial", size=9, color=MED_GRAY)
    ws7.merge_cells(f"A{rn+1+i}:C{rn+1+i}")

# Save
output = "ai-wealth-management-research.xlsx"
wb.save(output)
size_kb = os.path.getsize(output) / 1024
print(f"Created: {output} ({size_kb:.1f} KB)")
print(f"Sheets: {', '.join(wb.sheetnames)}")
