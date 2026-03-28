"""
AI Wealth Management Analysis — Comprehensive Excel Workbook
5 sheets: Executive Dashboard, Competitive Landscape, Regulatory Risk Matrix,
Investment Business Case, Implementation Timeline
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers, NamedStyle
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, DataBarRule, ColorScaleRule, FormulaRule
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins, PrintPageSetup
from openpyxl.workbook.defined_name import DefinedName
import os
from datetime import date, timedelta

wb = Workbook()

# ── Color palette ──
NAVY = "0B2545"
DARK_BLUE = "134074"
MID_BLUE = "13678A"
TEAL = "1B9AAA"
WHITE = "FFFFFF"
LIGHT_GRAY = "F8F9FA"
ALT_ROW = "EBF5FB"
GREEN_BG = "D5F5E3"
GREEN_FG = "1E8449"
YELLOW_BG = "FEF9E7"
YELLOW_FG = "B7950B"
RED_BG = "FADBD8"
RED_FG = "CB4335"
BLUE_INPUT = "0000FF"
BLACK = "000000"
GREEN_LINK = "008000"

# ── Reusable styles ──
hdr_font = Font(name="Arial", bold=True, color=WHITE, size=11)
hdr_fill = PatternFill("solid", fgColor=NAVY)
hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
body_font = Font(name="Arial", size=10)
body_align = Alignment(vertical="center", wrap_text=True)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
curr_fmt = '$#,##0'
curr_fmt_m = '$#,##0"M"'
pct_fmt = '0.0%'
date_fmt = 'MM/DD/YYYY'
thin = Side(style="thin", color="CBD5E1")
thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
dbl_side = Side(style="double", color=NAVY)
dbl_border = Border(top=dbl_side, bottom=dbl_side, left=thin, right=thin)


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = thin_border


def style_row(ws, row, ncols, alt=False):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = body_font
        cell.alignment = body_align
        cell.border = thin_border
        if alt:
            cell.fill = PatternFill("solid", fgColor=ALT_ROW)


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def set_print_settings(ws):
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5,
                                   header=0.3, footer=0.3)
    ws.sheet_properties.pageSetUpPr = None
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.oddFooter.center.text = "Confidential — AI Strategy Division"
    ws.oddFooter.right.text = "Page &P of &N"
    ws.print_title_rows = '1:2'


# ╔══════════════════════════════════════════════════════════════╗
# ║  SHEET 2 — Competitive Landscape (build first for refs)     ║
# ╚══════════════════════════════════════════════════════════════╝
ws2 = wb.active
ws2.title = "Competitive Landscape"

headers2 = [
    "Firm Name", "AUM ($B)", "AI Maturity (1-5)", "Annual AI Spend ($M)",
    "Key AI Initiatives", "Client-Facing AI", "Robo-Advisory", "AI Headcount"
]
widths2 = [22, 12, 16, 18, 45, 16, 15, 14]
set_col_widths(ws2, widths2)

for i, h in enumerate(headers2, 1):
    ws2.cell(row=1, column=i, value=h)
style_header(ws2, 1, len(headers2))

firms = [
    ["Morgan Stanley", 4600, 4.5, 1200, "AI @ Morgan Stanley (GPT-4), Next Best Action, Debrief AI", "Yes", "Yes", 1500],
    ["Bank of America", 3500, 4.3, 1100, "Erica Virtual Assistant (2B+ interactions), CashPro AI", "Yes", "Yes", 1200],
    ["Charles Schwab", 8500, 3.8, 650, "Schwab Intelligent Portfolios, Schwab Assistant AI", "Yes", "Yes", 800],
    ["Fidelity Investments", 4500, 4.0, 900, "Fidelity AMP, Fidelity Go, Signal AI Research", "Yes", "Yes", 1000],
    ["Vanguard", 8600, 3.5, 550, "Vanguard Digital Advisor ($365B AUM), PAS AI", "Yes", "Yes", 600],
    ["JP Morgan", 3900, 5.0, 1800, "IndexGPT, LLM Suite (60K employees), COiN, Jade AI", "Yes", "Yes", 2000],
    ["Goldman Sachs", 2800, 4.2, 950, "GS AI Platform, Marcus AI, Transaction Banking AI", "Yes", "No", 1100],
    ["UBS", 4000, 4.0, 800, "UBS Companion (Red AI), My Way advisory AI", "Yes", "Yes", 900],
    ["Wells Fargo", 2100, 3.5, 600, "Fargo AI Virtual Assistant, Predictive Banking AI", "Yes", "No", 700],
    ["Raymond James", 1300, 2.8, 180, "Client Insights AI, Portfolio Analytics ML", "Yes", "No", 200],
    ["Crestview Finance Group", 1900, 2.5, 150, "Advisor AI Assistant, Client Match ML", "No", "No", 150],
    ["LPL Financial", 1400, 3.0, 250, "ClientWorks AI, Compliance AI Monitoring", "Yes", "No", 300],
    ["Wealthfront", 50, 4.5, 80, "Automated Tax-Loss Harvesting, Risk Parity AI, Path Planner", "Yes", "Yes", 120],
    ["Betterment", 40, 4.3, 65, "Tax-Coordinated Portfolio, RetireGuide AI, Smart Rebalancing", "Yes", "Yes", 100],
    ["Robinhood", 130, 3.8, 200, "Cortex AI Recommendations, 24/7 AI Support, Robinhood Strategies", "Yes", "Yes", 350],
]

for r, row_data in enumerate(firms, 2):
    for c, val in enumerate(row_data, 1):
        cell = ws2.cell(row=r, column=c, value=val)
    ws2.cell(row=r, column=2).number_format = '#,##0'
    ws2.cell(row=r, column=4).number_format = '$#,##0'
    ws2.cell(row=r, column=8).number_format = '#,##0'
    style_row(ws2, r, len(headers2), alt=(r % 2 == 0))

# Data validation for Yes/No columns
yn_val = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
yn_val.error = "Please select Yes or No"
yn_val.errorTitle = "Invalid Entry"
ws2.add_data_validation(yn_val)
yn_val.add(f"F2:F{len(firms)+1}")
yn_val.add(f"G2:G{len(firms)+1}")

# Summary row
sr = len(firms) + 2
labels = ["SUMMARY", "", "", "", "", "", "", ""]
for c, lbl in enumerate(labels, 1):
    ws2.cell(row=sr, column=c, value=lbl if lbl else None)
ws2.cell(row=sr, column=1, value="SUMMARY (AVG / MAX / MIN)")
ws2.cell(row=sr, column=1).font = Font(name="Arial", bold=True, size=10)
ws2.cell(row=sr, column=2, value=f"=AVERAGE(B2:B{sr-1})")
ws2.cell(row=sr, column=2).number_format = '#,##0'
ws2.cell(row=sr, column=3, value=f"=AVERAGE(C2:C{sr-1})")
ws2.cell(row=sr, column=3).number_format = '0.0'
ws2.cell(row=sr, column=4, value=f"=SUM(D2:D{sr-1})")
ws2.cell(row=sr, column=4).number_format = '$#,##0'
ws2.cell(row=sr, column=8, value=f"=SUM(H2:H{sr-1})")
ws2.cell(row=sr, column=8).number_format = '#,##0'
for c in range(1, len(headers2) + 1):
    ws2.cell(row=sr, column=c).border = dbl_border
    ws2.cell(row=sr, column=c).font = Font(name="Arial", bold=True, size=10)

# Max/Min rows
for label, func, offset in [("MAX", "MAX", 1), ("MIN", "MIN", 2)]:
    row = sr + offset
    ws2.cell(row=row, column=1, value=label)
    ws2.cell(row=row, column=1).font = Font(name="Arial", bold=True, size=10)
    ws2.cell(row=row, column=2, value=f"={func}(B2:B{sr-1})")
    ws2.cell(row=row, column=2).number_format = '#,##0'
    ws2.cell(row=row, column=3, value=f"={func}(C2:C{sr-1})")
    ws2.cell(row=row, column=3).number_format = '0.0'
    ws2.cell(row=row, column=4, value=f"={func}(D2:D{sr-1})")
    ws2.cell(row=row, column=4).number_format = '$#,##0'
    ws2.cell(row=row, column=8, value=f"={func}(H2:H{sr-1})")
    ws2.cell(row=row, column=8).number_format = '#,##0'
    for c in range(1, len(headers2) + 1):
        ws2.cell(row=row, column=c).border = dbl_border

# Data bars on AI Spend (col D)
ws2.conditional_formatting.add(
    f"D2:D{len(firms)+1}",
    DataBarRule(start_type="min", end_type="max", color="3498DB")
)

# Color scale on AI Maturity (col C)
ws2.conditional_formatting.add(
    f"C2:C{len(firms)+1}",
    ColorScaleRule(
        start_type="min", start_color="F1948A",
        mid_type="percentile", mid_value=50, mid_color="F9E79F",
        end_type="max", end_color="82E0AA"
    )
)

ws2.auto_filter.ref = f"A1:H{len(firms)+1}"
ws2.freeze_panes = "A2"
set_print_settings(ws2)


# ╔══════════════════════════════════════════════════════════════╗
# ║  SHEET 3 — Regulatory Risk Matrix                          ║
# ╚══════════════════════════════════════════════════════════════╝
ws3 = wb.create_sheet("Regulatory Risk Matrix")

headers3 = [
    "Regulation / Guidance", "Issuing Body", "Effective Date",
    "Risk Level", "Impact Area", "Compliance Status",
    "Action Required", "Deadline", "Owner"
]
widths3 = [32, 14, 14, 12, 18, 16, 38, 14, 16]
set_col_widths(ws3, widths3)

# COUNTIF summary at top (rows 1-2)
ws3.merge_cells("A1:B1")
ws3.cell(row=1, column=1, value="Risk Summary")
ws3.cell(row=1, column=1).font = Font(name="Arial", bold=True, size=14, color=NAVY)

risk_labels = ["Critical", "High", "Medium", "Low"]
for i, rl in enumerate(risk_labels):
    col = 3 + i
    ws3.cell(row=1, column=col, value=rl)
    ws3.cell(row=1, column=col).font = Font(name="Arial", bold=True, size=10, color=WHITE)
    colors = {"Critical": RED_FG, "High": YELLOW_FG, "Medium": MID_BLUE, "Low": GREEN_FG}
    ws3.cell(row=1, column=col).fill = PatternFill("solid", fgColor=colors[rl].replace("#", ""))
    ws3.cell(row=1, column=col).alignment = center_align
    ws3.cell(row=2, column=col, value=f'=COUNTIF(D5:D50,"{rl}")')
    ws3.cell(row=2, column=col).font = Font(name="Georgia", bold=True, size=18, color=NAVY)
    ws3.cell(row=2, column=col).alignment = center_align

# Header row at row 4
hdr_row = 4
for i, h in enumerate(headers3, 1):
    ws3.cell(row=hdr_row, column=i, value=h)
style_header(ws3, hdr_row, len(headers3))

regs = [
    ["SEC Proposed PDA/AI Conflicts Rule (Reg. 23-3)", "SEC", "2025-06-15", "Critical",
     "Model Risk / Conflicts", "In Progress",
     "Implement conflict-of-interest elimination for AI/PDA in advisory", "2025-09-30", "Chief Compliance Officer"],
    ["SR 11-7 Model Risk Management", "Federal Reserve", "2011-04-04", "Critical",
     "Model Risk", "In Progress",
     "Full MRM framework for all AI/ML models: validation, documentation, monitoring", "2025-06-30", "Model Risk Mgmt"],
    ["FINRA Report on AI (RN 24-09)", "FINRA", "2024-12-01", "High",
     "Supervisory / Suitability", "In Progress",
     "Supervisory procedures for AI-driven recommendations, Reg BI alignment", "2025-12-31", "Compliance Director"],
    ["SEC Marketing Rule — AI-Washing", "SEC", "2022-11-04", "High",
     "Disclosure", "In Progress",
     "Review all AI marketing claims for substantiation; remove misleading AI references", "2025-06-30", "Marketing/Legal"],
    ["OCC Third-Party Risk Mgmt (2023-17)", "OCC", "2023-06-06", "High",
     "Vendor Risk", "In Progress",
     "Due diligence on AI vendors, concentration risk assessment, exit strategies", "2025-09-30", "Vendor Mgmt"],
    ["CFPB Circular 2022-03 (Adverse Action AI)", "CFPB", "2022-05-26", "Critical",
     "Fair Lending / Consumer", "Gap",
     "Ensure AI credit decisions provide specific adverse action reasons, not generic", "2025-06-30", "Fair Lending Officer"],
    ["ECOA / Reg B Fair Lending (AI Bias)", "CFPB / DOJ", "1974-10-28", "Critical",
     "Fair Lending", "In Progress",
     "Disparate impact testing for all AI models, bias audits, alternative data review", "2025-09-30", "Fair Lending Officer"],
    ["GLBA (Gramm-Leach-Bliley Act)", "FTC / Regulators", "1999-11-12", "High",
     "Data Privacy", "Compliant",
     "Privacy notices for AI data usage, opt-out rights, safeguard AI training data", "Ongoing", "Privacy Officer"],
    ["CCPA / CPRA (California)", "CA Attorney General", "2023-01-01", "High",
     "Data Privacy", "In Progress",
     "AI profiling opt-out, automated decision-making disclosure, data deletion rights", "2025-12-31", "Privacy Officer"],
    ["NIST AI Risk Management Framework 1.0", "NIST", "2023-01-26", "Medium",
     "Governance", "In Progress",
     "Map AI systems to NIST AI RMF categories; implement Govern/Map/Measure/Manage", "2026-06-30", "AI Governance Lead"],
    ["EU AI Act (Extraterritorial)", "European Commission", "2025-08-02", "Medium",
     "Governance / Classification", "Gap",
     "Classify AI systems by risk tier; high-risk financial AI needs conformity assessment", "2026-08-01", "Global Compliance"],
    ["Colorado SB 24-205 (AI Consumer Protection)", "State of Colorado", "2026-02-01", "Medium",
     "Consumer Protection", "Gap",
     "Impact assessments for high-risk AI, consumer notification, algorithmic fairness", "2026-02-01", "State Compliance"],
]

for r, row_data in enumerate(regs, hdr_row + 1):
    for c, val in enumerate(row_data, 1):
        ws3.cell(row=r, column=c, value=val)
    if row_data[2] != "Ongoing":
        ws3.cell(row=r, column=3).number_format = date_fmt
    if row_data[7] != "Ongoing":
        ws3.cell(row=r, column=8).number_format = date_fmt
    style_row(ws3, r, len(headers3), alt=(r % 2 == 0))

# Data validation for Compliance Status
comp_val = DataValidation(
    type="list",
    formula1='"Compliant,In Progress,Gap,Not Applicable"',
    allow_blank=True
)
ws3.add_data_validation(comp_val)
comp_val.add(f"F{hdr_row+1}:F{hdr_row+len(regs)}")

# Conditional formatting — Risk Level
ws3.conditional_formatting.add(
    f"D{hdr_row+1}:D{hdr_row+50}",
    CellIsRule(operator="equal", formula=['"Critical"'],
              fill=PatternFill("solid", fgColor=RED_BG.replace("#", "")),
              font=Font(color=RED_FG, bold=True))
)
ws3.conditional_formatting.add(
    f"D{hdr_row+1}:D{hdr_row+50}",
    CellIsRule(operator="equal", formula=['"High"'],
              fill=PatternFill("solid", fgColor=YELLOW_BG.replace("#", "")),
              font=Font(color=YELLOW_FG, bold=True))
)
ws3.conditional_formatting.add(
    f"D{hdr_row+1}:D{hdr_row+50}",
    CellIsRule(operator="equal", formula=['"Low"'],
              fill=PatternFill("solid", fgColor=GREEN_BG.replace("#", "")),
              font=Font(color=GREEN_FG, bold=True))
)
ws3.conditional_formatting.add(
    f"D{hdr_row+1}:D{hdr_row+50}",
    CellIsRule(operator="equal", formula=['"Medium"'],
              fill=PatternFill("solid", fgColor="D6EAF8"),
              font=Font(color=MID_BLUE, bold=True))
)

ws3.freeze_panes = f"A{hdr_row+1}"
set_print_settings(ws3)


# ╔══════════════════════════════════════════════════════════════╗
# ║  SHEET 4 — Investment Business Case                        ║
# ╚══════════════════════════════════════════════════════════════╝
ws4 = wb.create_sheet("Investment Business Case")

# Title
ws4.merge_cells("A1:F1")
ws4.cell(row=1, column=1, value="AI Investment Business Case — FY 2026")
ws4.cell(row=1, column=1).font = Font(name="Georgia", bold=True, size=14, color=NAVY)

# Key Assumptions section (rows 2-4)
ws4.cell(row=2, column=1, value="Key Assumptions:")
ws4.cell(row=2, column=1).font = Font(name="Arial", bold=True, size=10, color=NAVY)

assumptions = [
    ("Discount Rate", 0.10, "B3", "discount_rate"),
    ("Growth Rate", 0.15, "D3", "growth_rate"),
    ("Avg Headcount Cost ($K)", 185, "F3", "headcount_cost"),
]
for label, val, cell_ref, name in assumptions:
    col_letter = cell_ref[0]
    row_num = int(cell_ref[1:])
    label_col = ord(col_letter) - 64 - 1
    val_col = ord(col_letter) - 64
    ws4.cell(row=row_num, column=label_col, value=label)
    ws4.cell(row=row_num, column=label_col).font = Font(name="Arial", size=9, color="666666")
    ws4.cell(row=row_num, column=val_col, value=val)
    ws4.cell(row=row_num, column=val_col).font = Font(name="Arial", bold=True, size=10, color=BLUE_INPUT)
    if "Rate" in label:
        ws4.cell(row=row_num, column=val_col).number_format = '0.0%'
    else:
        ws4.cell(row=row_num, column=val_col).number_format = '$#,##0'

# Define named ranges
for label, val, cell_ref, name in assumptions:
    dn = DefinedName(name, attr_text=f"'Investment Business Case'!${cell_ref[0]}${cell_ref[1:]}")
    wb.defined_names.add(dn)

# Column headers (row 5)
col_headers = ["Investment Category", "Q1 2026", "Q2 2026", "Q3 2026", "Q4 2026", "FY 2026 Total"]
widths4 = [30, 16, 16, 16, 16, 18]
set_col_widths(ws4, widths4)

for i, h in enumerate(col_headers, 1):
    ws4.cell(row=5, column=i, value=h)
style_header(ws4, 5, len(col_headers))

# Investment categories with quarterly input values (blue = input)
categories = [
    ("AI Platform Licensing", [450, 450, 475, 475]),
    ("Data Infrastructure", [380, 400, 420, 440]),
    ("ML Engineering Headcount", [925, 960, 1000, 1050]),
    ("Training & Change Management", [120, 150, 100, 80]),
    ("Compliance & Audit", [200, 180, 190, 210]),
    ("Vendor Integration", [160, 175, 150, 140]),
]

for r, (cat, qtrs) in enumerate(categories, 6):
    ws4.cell(row=r, column=1, value=cat)
    ws4.cell(row=r, column=1).font = Font(name="Arial", size=10, color=BLACK)
    for q, val in enumerate(qtrs, 2):
        cell = ws4.cell(row=r, column=q, value=val)
        cell.font = Font(name="Arial", size=10, color=BLUE_INPUT)
        cell.number_format = curr_fmt
    # FY Total formula
    ws4.cell(row=r, column=6, value=f"=SUM(B{r}:E{r})")
    ws4.cell(row=r, column=6).font = Font(name="Arial", bold=True, size=10, color=BLACK)
    ws4.cell(row=r, column=6).number_format = curr_fmt
    style_row(ws4, r, 6, alt=(r % 2 == 0))

# Total Investment row
total_r = 6 + len(categories)
ws4.cell(row=total_r, column=1, value="TOTAL INVESTMENT ($K)")
ws4.cell(row=total_r, column=1).font = Font(name="Arial", bold=True, size=11, color=NAVY)
for c in range(2, 7):
    col_letter = get_column_letter(c)
    ws4.cell(row=total_r, column=c, value=f"=SUM({col_letter}6:{col_letter}{total_r-1})")
    ws4.cell(row=total_r, column=c).font = Font(name="Arial", bold=True, size=11, color=BLACK)
    ws4.cell(row=total_r, column=c).number_format = curr_fmt
    ws4.cell(row=total_r, column=c).border = dbl_border

# Budget row
budget_r = total_r + 1
ws4.cell(row=budget_r, column=1, value="BUDGET ($K)")
ws4.cell(row=budget_r, column=1).font = Font(name="Arial", bold=True, size=10, color=BLUE_INPUT)
budgets = [2300, 2400, 2400, 2500]
for c, val in enumerate(budgets, 2):
    ws4.cell(row=budget_r, column=c, value=val)
    ws4.cell(row=budget_r, column=c).font = Font(name="Arial", size=10, color=BLUE_INPUT)
    ws4.cell(row=budget_r, column=c).number_format = curr_fmt
ws4.cell(row=budget_r, column=6, value=f"=SUM(B{budget_r}:E{budget_r})")
ws4.cell(row=budget_r, column=6).font = Font(name="Arial", bold=True, size=10, color=BLACK)
ws4.cell(row=budget_r, column=6).number_format = curr_fmt
style_row(ws4, budget_r, 6)

# Variance row
var_r = budget_r + 1
ws4.cell(row=var_r, column=1, value="VARIANCE (Projected - Budget)")
ws4.cell(row=var_r, column=1).font = Font(name="Arial", bold=True, size=10, color=NAVY)
for c in range(2, 7):
    col_l = get_column_letter(c)
    ws4.cell(row=var_r, column=c, value=f"={col_l}{total_r}-{col_l}{budget_r}")
    ws4.cell(row=var_r, column=c).font = Font(name="Arial", bold=True, size=10, color=BLACK)
    ws4.cell(row=var_r, column=c).number_format = '$#,##0;[Red]($#,##0);"-"'

# Conditional formatting: red if over budget (variance > 0)
ws4.conditional_formatting.add(
    f"B{var_r}:F{var_r}",
    CellIsRule(operator="greaterThan", formula=["0"],
              fill=PatternFill("solid", fgColor=RED_BG),
              font=Font(color=RED_FG, bold=True))
)
ws4.conditional_formatting.add(
    f"B{var_r}:F{var_r}",
    CellIsRule(operator="lessThanOrEqual", formula=["0"],
              fill=PatternFill("solid", fgColor=GREEN_BG),
              font=Font(color=GREEN_FG, bold=True))
)

# ROI section
roi_r = var_r + 2
ws4.cell(row=roi_r, column=1, value="PROJECTED BENEFITS ($K)")
ws4.cell(row=roi_r, column=1).font = Font(name="Arial", bold=True, size=11, color=NAVY)
benefits = [
    ("Operational Efficiency Savings", [320, 480, 650, 820]),
    ("Revenue Uplift (AI-driven AUM growth)", [150, 300, 500, 750]),
    ("Compliance Cost Avoidance", [80, 100, 120, 140]),
    ("Client Retention Value", [100, 180, 260, 350]),
]
for r, (cat, qtrs) in enumerate(benefits, roi_r + 1):
    ws4.cell(row=r, column=1, value=cat)
    for q, val in enumerate(qtrs, 2):
        ws4.cell(row=r, column=q, value=val)
        ws4.cell(row=r, column=q).font = Font(name="Arial", size=10, color=BLUE_INPUT)
        ws4.cell(row=r, column=q).number_format = curr_fmt
    ws4.cell(row=r, column=6, value=f"=SUM(B{r}:E{r})")
    ws4.cell(row=r, column=6).font = Font(name="Arial", bold=True, size=10)
    ws4.cell(row=r, column=6).number_format = curr_fmt
    style_row(ws4, r, 6, alt=(r % 2 == 0))

ben_total_r = roi_r + 1 + len(benefits)
ws4.cell(row=ben_total_r, column=1, value="TOTAL BENEFITS ($K)")
ws4.cell(row=ben_total_r, column=1).font = Font(name="Arial", bold=True, size=11, color=NAVY)
for c in range(2, 7):
    col_l = get_column_letter(c)
    ws4.cell(row=ben_total_r, column=c,
             value=f"=SUM({col_l}{roi_r+1}:{col_l}{ben_total_r-1})")
    ws4.cell(row=ben_total_r, column=c).font = Font(name="Arial", bold=True, size=11)
    ws4.cell(row=ben_total_r, column=c).number_format = curr_fmt
    ws4.cell(row=ben_total_r, column=c).border = dbl_border

# Net Value & ROI
net_r = ben_total_r + 1
ws4.cell(row=net_r, column=1, value="NET VALUE ($K)")
ws4.cell(row=net_r, column=1).font = Font(name="Arial", bold=True, size=11, color=NAVY)
for c in range(2, 7):
    col_l = get_column_letter(c)
    ws4.cell(row=net_r, column=c, value=f"={col_l}{ben_total_r}-{col_l}{total_r}")
    ws4.cell(row=net_r, column=c).font = Font(name="Arial", bold=True, size=11)
    ws4.cell(row=net_r, column=c).number_format = '$#,##0;[Red]($#,##0)'

roi_calc_r = net_r + 1
ws4.cell(row=roi_calc_r, column=1, value="ROI %")
ws4.cell(row=roi_calc_r, column=1).font = Font(name="Arial", bold=True, size=11, color=NAVY)
for c in range(2, 7):
    col_l = get_column_letter(c)
    ws4.cell(row=roi_calc_r, column=c,
             value=f"=IF({col_l}{total_r}=0,0,{col_l}{net_r}/{col_l}{total_r})")
    ws4.cell(row=roi_calc_r, column=c).font = Font(name="Arial", bold=True, size=11)
    ws4.cell(row=roi_calc_r, column=c).number_format = '0.0%'

ws4.freeze_panes = "A6"
set_print_settings(ws4)


# ╔══════════════════════════════════════════════════════════════╗
# ║  SHEET 5 — Implementation Timeline                         ║
# ╚══════════════════════════════════════════════════════════════╝
ws5 = wb.create_sheet("Implementation Timeline")

headers5 = [
    "Phase", "Initiative", "Start Date", "End Date",
    "Duration (Days)", "Status", "Dependencies", "Owner", "Budget ($K)"
]
widths5 = [14, 32, 14, 14, 14, 14, 24, 18, 14]
set_col_widths(ws5, widths5)

for i, h in enumerate(headers5, 1):
    ws5.cell(row=1, column=i, value=h)
style_header(ws5, 1, len(headers5))

base = date(2026, 4, 1)
timeline = [
    ["Phase 1", "AI Governance Framework Setup", base, base + timedelta(days=60),
     None, "In Progress", "None", "AI Governance Lead", 150],
    ["Phase 1", "Data Infrastructure Assessment", base + timedelta(days=14), base + timedelta(days=75),
     None, "In Progress", "None", "CTO Office", 280],
    ["Phase 1", "Vendor Selection & Procurement", base + timedelta(days=30), base + timedelta(days=105),
     None, "Not Started", "Governance Framework", "Vendor Mgmt", 120],
    ["Phase 2", "ML Platform Deployment", base + timedelta(days=90), base + timedelta(days=180),
     None, "Not Started", "Data Infrastructure", "ML Engineering", 650],
    ["Phase 2", "Model Risk Management Implementation", base + timedelta(days=90), base + timedelta(days=195),
     None, "Not Started", "Governance Framework", "Model Risk Mgmt", 320],
    ["Phase 2", "Client Sentiment Analysis MVP", base + timedelta(days=105), base + timedelta(days=195),
     None, "Not Started", "ML Platform", "Data Science", 180],
    ["Phase 2", "Compliance AI Monitoring Pilot", base + timedelta(days=120), base + timedelta(days=210),
     None, "Not Started", "ML Platform, MRM", "Compliance", 250],
    ["Phase 3", "AI-Assisted Portfolio Rebalancing", base + timedelta(days=180), base + timedelta(days=300),
     None, "Not Started", "ML Platform, MRM", "Portfolio Mgmt", 420],
    ["Phase 3", "NLP Client Communications", base + timedelta(days=195), base + timedelta(days=315),
     None, "Not Started", "Sentiment Analysis", "Client Services", 290],
    ["Phase 3", "Predictive Client Churn Model", base + timedelta(days=210), base + timedelta(days=330),
     None, "Not Started", "ML Platform", "Data Science", 200],
    ["Phase 3", "Advisor AI Assistant Rollout", base + timedelta(days=240), base + timedelta(days=360),
     None, "Not Started", "NLP, Portfolio AI", "Advisory", 380],
    ["Phase 4", "Robo-Advisory Enhancement", base + timedelta(days=300), base + timedelta(days=420),
     None, "Not Started", "Portfolio Rebalancing", "Digital Advisory", 350],
    ["Phase 4", "Full Production AI Monitoring", base + timedelta(days=330), base + timedelta(days=450),
     None, "Not Started", "All Phase 3", "AI Governance", 180],
    ["Phase 4", "Client-Facing AI Chatbot Launch", base + timedelta(days=360), base + timedelta(days=480),
     None, "Not Started", "NLP, Compliance AI", "Digital", 260],
    ["Phase 4", "AI ROI Assessment & Optimization", base + timedelta(days=450), base + timedelta(days=540),
     None, "Not Started", "All initiatives", "CFO / Strategy", 90],
]

for r, row_data in enumerate(timeline, 2):
    for c, val in enumerate(row_data, 1):
        if c == 5:  # Duration formula
            ws5.cell(row=r, column=c, value=f"=D{r}-C{r}")
        else:
            ws5.cell(row=r, column=c, value=val)
    ws5.cell(row=r, column=3).number_format = date_fmt
    ws5.cell(row=r, column=4).number_format = date_fmt
    ws5.cell(row=r, column=5).number_format = '#,##0'
    ws5.cell(row=r, column=9).number_format = curr_fmt
    style_row(ws5, r, len(headers5), alt=(r % 2 == 0))

# Status dropdown
status_val = DataValidation(
    type="list",
    formula1='"Not Started,In Progress,Complete,Blocked"',
    allow_blank=True
)
ws5.add_data_validation(status_val)
status_val.add(f"F2:F{len(timeline)+1}")

# Status conditional formatting
ws5.conditional_formatting.add(
    f"F2:F{len(timeline)+50}",
    CellIsRule(operator="equal", formula=['"Complete"'],
              fill=PatternFill("solid", fgColor=GREEN_BG),
              font=Font(color=GREEN_FG, bold=True))
)
ws5.conditional_formatting.add(
    f"F2:F{len(timeline)+50}",
    CellIsRule(operator="equal", formula=['"In Progress"'],
              fill=PatternFill("solid", fgColor=YELLOW_BG),
              font=Font(color=YELLOW_FG, bold=True))
)
ws5.conditional_formatting.add(
    f"F2:F{len(timeline)+50}",
    CellIsRule(operator="equal", formula=['"Blocked"'],
              fill=PatternFill("solid", fgColor=RED_BG),
              font=Font(color=RED_FG, bold=True))
)

# Total budget row
bt_r = len(timeline) + 2
ws5.cell(row=bt_r, column=1, value="TOTAL")
ws5.cell(row=bt_r, column=1).font = Font(name="Arial", bold=True, size=11)
ws5.cell(row=bt_r, column=9, value=f"=SUM(I2:I{bt_r-1})")
ws5.cell(row=bt_r, column=9).font = Font(name="Arial", bold=True, size=11)
ws5.cell(row=bt_r, column=9).number_format = curr_fmt
for c in range(1, len(headers5) + 1):
    ws5.cell(row=bt_r, column=c).border = dbl_border

# Cumulative Budget Line Chart
# Build cumulative budget data below the table
cum_start = bt_r + 3
ws5.cell(row=cum_start, column=1, value="Initiative")
ws5.cell(row=cum_start, column=2, value="Cumulative Budget ($K)")
style_header(ws5, cum_start, 2)

for r_idx, tl_row in enumerate(timeline, cum_start + 1):
    ws5.cell(row=r_idx, column=1, value=tl_row[1])
    if r_idx == cum_start + 1:
        ws5.cell(row=r_idx, column=2, value=f"=I2")
    else:
        orig_data_row = r_idx - cum_start
        ws5.cell(row=r_idx, column=2, value=f"=B{r_idx-1}+I{orig_data_row + 1}")
    ws5.cell(row=r_idx, column=2).number_format = curr_fmt

chart_line = LineChart()
chart_line.title = "Cumulative AI Investment Budget ($K)"
chart_line.style = 10
chart_line.y_axis.title = "Cumulative Spend ($K)"
chart_line.x_axis.title = "Initiative"
chart_line.height = 15
chart_line.width = 28

data_ref = Reference(ws5, min_col=2, min_row=cum_start,
                     max_row=cum_start + len(timeline))
cats_ref = Reference(ws5, min_col=1, min_row=cum_start + 1,
                     max_row=cum_start + len(timeline))
chart_line.add_data(data_ref, titles_from_data=True)
chart_line.set_categories(cats_ref)
chart_line.series[0].graphicalProperties.line.width = 25000
ws5.add_chart(chart_line, f"A{cum_start + len(timeline) + 2}")

ws5.freeze_panes = "A2"
set_print_settings(ws5)


# ╔══════════════════════════════════════════════════════════════╗
# ║  SHEET 1 — Executive Dashboard (last, so it can reference)  ║
# ╚══════════════════════════════════════════════════════════════╝
ws1 = wb.create_sheet("Executive Dashboard")
wb.move_sheet("Executive Dashboard", offset=-4)

# Title row (merged)
ws1.merge_cells("A1:H1")
ws1.cell(row=1, column=1, value="AI in Wealth Management — Executive Dashboard")
ws1.cell(row=1, column=1).font = Font(name="Georgia", bold=True, size=18, color=WHITE)
ws1.cell(row=1, column=1).fill = PatternFill("solid", fgColor=NAVY)
ws1.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 45

ws1.merge_cells("A2:H2")
ws1.cell(row=2, column=1, value="Prepared: March 2026  |  Confidential — AI Strategy Division")
ws1.cell(row=2, column=1).font = Font(name="Arial", size=10, color=WHITE, italic=True)
ws1.cell(row=2, column=1).fill = PatternFill("solid", fgColor=DARK_BLUE)
ws1.cell(row=2, column=1).alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[2].height = 25

set_col_widths(ws1, [18, 16, 16, 16, 18, 16, 16, 16])

# KPI Cards (row 4-5)
ws1.cell(row=3, column=1, value="KEY PERFORMANCE INDICATORS")
ws1.cell(row=3, column=1).font = Font(name="Arial", bold=True, size=12, color=NAVY)
ws1.row_dimensions[4].height = 20
ws1.row_dimensions[5].height = 45

kpis = [
    ("Total AI Market Size", '="$3.7B"', "A", None),
    ("AI Adoption Rate", "=80/100", "B", pct_fmt),
    ("Avg AI Spend/Firm", "=AVERAGE('Competitive Landscape'!D2:D16)", "C", '$#,##0"M"'),
    ("Regulatory Risk Score", '=COUNTIF(\'Regulatory Risk Matrix\'!D5:D16,"Critical")&"/"&COUNTA(\'Regulatory Risk Matrix\'!D5:D16)', "D", None),
    ("Projected 3-Year ROI", "='Investment Business Case'!F" + str(roi_calc_r), "E", '0.0%'),
    ("Client Satisfaction Δ", "=0.12", "F", '+0.0%;-0.0%'),
]

for i, (label, formula, col, fmt) in enumerate(kpis):
    c = i + 1
    # Label
    ws1.cell(row=4, column=c, value=label)
    ws1.cell(row=4, column=c).font = Font(name="Arial", size=9, color="64748B")
    ws1.cell(row=4, column=c).alignment = center_align
    ws1.cell(row=4, column=c).fill = PatternFill("solid", fgColor=LIGHT_GRAY)
    ws1.cell(row=4, column=c).border = thin_border
    # Value
    ws1.cell(row=5, column=c, value=formula)
    ws1.cell(row=5, column=c).font = Font(name="Georgia", bold=True, size=20, color=NAVY)
    ws1.cell(row=5, column=c).alignment = center_align
    ws1.cell(row=5, column=c).border = thin_border
    if fmt:
        ws1.cell(row=5, column=c).number_format = fmt

# Bar Chart — AI Investment by Firm (top 8)
ws1.cell(row=7, column=1, value="AI Investment by Firm (Top 8, $M)")
ws1.cell(row=7, column=1).font = Font(name="Arial", bold=True, size=12, color=NAVY)

chart_bar = BarChart()
chart_bar.type = "col"
chart_bar.title = "Annual AI Spend by Firm ($M)"
chart_bar.style = 10
chart_bar.y_axis.title = "AI Spend ($M)"
chart_bar.height = 15
chart_bar.width = 22

# Reference top 8 firms from Competitive Landscape (sorted by spend — JP Morgan is row 7 etc.)
# Use inline data reference: rows 2-9 from col A (names) and col D (spend)
data_bar = Reference(ws2, min_col=4, min_row=1, max_row=9)
cats_bar = Reference(ws2, min_col=1, min_row=2, max_row=9)
chart_bar.add_data(data_bar, titles_from_data=True)
chart_bar.set_categories(cats_bar)
chart_bar.series[0].graphicalProperties.solidFill = "3498DB"
ws1.add_chart(chart_bar, "A8")

# Pie Chart — AI Use Case Distribution
ws1.cell(row=7, column=5, value="AI Use Case Distribution")
ws1.cell(row=7, column=5).font = Font(name="Arial", bold=True, size=12, color=NAVY)

# Pie data (embedded in dashboard)
pie_data_start = 24
pie_labels = ["Advisory & Planning", "Compliance & Risk", "Operations", "Client Service"]
pie_values = [35, 18, 15, 32]
ws1.cell(row=pie_data_start, column=5, value="Use Case")
ws1.cell(row=pie_data_start, column=6, value="% of AI Spend")
for i, (lbl, val) in enumerate(zip(pie_labels, pie_values)):
    ws1.cell(row=pie_data_start + 1 + i, column=5, value=lbl)
    ws1.cell(row=pie_data_start + 1 + i, column=6, value=val)

chart_pie = PieChart()
chart_pie.title = "AI Use Case Distribution (% of Spend)"
chart_pie.style = 10
chart_pie.height = 15
chart_pie.width = 16

pie_d = Reference(ws1, min_col=6, min_row=pie_data_start,
                  max_row=pie_data_start + len(pie_labels))
pie_c = Reference(ws1, min_col=5, min_row=pie_data_start + 1,
                  max_row=pie_data_start + len(pie_labels))
chart_pie.add_data(pie_d, titles_from_data=True)
chart_pie.set_categories(pie_c)

# Color the pie slices
colors_pie = ["3498DB", "E74C3C", "F39C12", "2ECC71"]
for idx, color in enumerate(colors_pie):
    pt = DataPoint(idx=idx)
    pt.graphicalProperties.solidFill = color
    chart_pie.series[0].data_points.append(pt)

chart_pie.dataLabels = DataLabelList()
chart_pie.dataLabels.showPercent = True
chart_pie.dataLabels.showCatName = True

ws1.add_chart(chart_pie, "E8")

ws1.freeze_panes = "A3"
set_print_settings(ws1)

# ── Save ──
output = "ai_wealth_management_analysis.xlsx"
wb.save(output)
size_kb = os.path.getsize(output) / 1024
print(f"Created: {output} ({size_kb:.1f} KB)")
print(f"Sheets: {', '.join(wb.sheetnames)}")
