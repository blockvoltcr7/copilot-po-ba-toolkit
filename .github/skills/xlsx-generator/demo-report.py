"""
Demo: AI Agent Skills Adoption Tracker
Generates a multi-sheet Excel workbook with styled headers,
formulas, conditional formatting, and frozen panes.

Usage: pip install openpyxl && python demo-report.py
Output: ai_skills_adoption_tracker.xlsx
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
import os

wb = Workbook()

# --- Color palette ---
DEEP_BLUE = "065A82"
TEAL = "1C7293"
MIDNIGHT = "21295C"
WHITE = "FFFFFF"
LIGHT_BG = "F0F4F8"
LIGHT_BLUE = "DBEAFE"
GREEN = "16A34A"
YELLOW = "EAB308"
RED = "DC2626"

# --- Reusable styles ---
header_font = Font(name="Arial", bold=True, color=WHITE, size=11)
header_fill = PatternFill("solid", fgColor=DEEP_BLUE)
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
body_font = Font(name="Arial", size=10)
body_align = Alignment(vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1"),
)
alt_fill = PatternFill("solid", fgColor=LIGHT_BG)

def style_header_row(ws, row, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

def style_data_row(ws, row, col_count, alternate=False):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = body_font
        cell.alignment = body_align
        cell.border = thin_border
        if alternate:
            cell.fill = alt_fill

# ============================================================
# SHEET 1 — Skills Inventory
# ============================================================
ws1 = wb.active
ws1.title = "Skills Inventory"

headers1 = ["Skill Name", "Category", "Tier", "Language", "Status", "Adoption %"]
col_widths1 = [28, 18, 14, 12, 14, 14]

for i, (header, width) in enumerate(zip(headers1, col_widths1), 1):
    ws1.cell(row=1, column=i, value=header)
    ws1.column_dimensions[get_column_letter(i)].width = width

style_header_row(ws1, 1, len(headers1))

skills_data = [
    ["docx-generator", "Document Gen", "Day 1", "Node.js", "Active", 0.85],
    ["pptx-generator", "Document Gen", "Day 1", "Node.js", "Active", 0.72],
    ["xlsx-generator", "Document Gen", "Day 1", "Python", "Active", 0.65],
    ["frontend-design", "Design", "Week 1", "TypeScript", "Active", 0.90],
    ["shadcn", "Design", "Week 1", "TypeScript", "Active", 0.88],
    ["browse", "QA/Testing", "Week 1", "Node.js", "Active", 0.78],
    ["qa", "QA/Testing", "Week 2-3", "Node.js", "Active", 0.60],
    ["investigate", "Debugging", "Week 1", "Node.js", "Active", 0.82],
    ["ship", "Deployment", "Week 2-3", "Bash", "Active", 0.55],
    ["review", "Code Review", "Week 2-3", "Node.js", "Active", 0.70],
    ["mcp-builder", "Integration", "Week 4+", "TypeScript", "Planned", 0.20],
    ["office-hours", "Strategy", "Week 4+", "Node.js", "Planned", 0.15],
    ["retro", "Team", "Week 4+", "Node.js", "Planned", 0.10],
    ["design-review", "Design", "Week 2-3", "Node.js", "Active", 0.68],
    ["vitest", "Testing", "Week 1", "TypeScript", "Active", 0.75],
]

for r, row_data in enumerate(skills_data, 2):
    for c, value in enumerate(row_data, 1):
        ws1.cell(row=r, column=c, value=value)
    # Format adoption % column
    ws1.cell(row=r, column=6).number_format = "0%"
    style_data_row(ws1, r, len(headers1), alternate=(r % 2 == 0))

# Conditional formatting on Status column (E)
ws1.conditional_formatting.add(
    "E2:E100",
    CellIsRule(operator="equal", formula=['"Active"'],
              fill=PatternFill("solid", fgColor="DCFCE7"),
              font=Font(color=GREEN))
)
ws1.conditional_formatting.add(
    "E2:E100",
    CellIsRule(operator="equal", formula=['"Planned"'],
              fill=PatternFill("solid", fgColor="FEF9C3"),
              font=Font(color=YELLOW))
)

# Freeze top row
ws1.freeze_panes = "A2"
# Auto-filter
ws1.auto_filter.ref = f"A1:F{len(skills_data) + 1}"

# ============================================================
# SHEET 2 — Summary Dashboard
# ============================================================
ws2 = wb.create_sheet("Summary Dashboard")

# Title
ws2.merge_cells("A1:D1")
ws2.cell(row=1, column=1, value="AI Skills Adoption Dashboard")
ws2.cell(row=1, column=1).font = Font(name="Georgia", bold=True, size=16, color=MIDNIGHT)
ws2.cell(row=1, column=1).alignment = Alignment(horizontal="left")
ws2.row_dimensions[1].height = 35

# KPI cards row
kpis = [
    ("Total Skills", f"=COUNTA('Skills Inventory'!A2:A100)"),
    ("Active Skills", f"=COUNTIF('Skills Inventory'!E2:E100,\"Active\")"),
    ("Avg Adoption", f"=AVERAGE('Skills Inventory'!F2:F{len(skills_data)+1})"),
    ("Day 1 Skills", f"=COUNTIF('Skills Inventory'!C2:C100,\"Day 1\")"),
]

for i, (label, formula) in enumerate(kpis):
    col = i + 1
    ws2.column_dimensions[get_column_letter(col)].width = 20

    # Label
    ws2.cell(row=3, column=col, value=label)
    ws2.cell(row=3, column=col).font = Font(name="Arial", size=10, color="64748B")
    ws2.cell(row=3, column=col).alignment = Alignment(horizontal="center")

    # Value (formula)
    ws2.cell(row=4, column=col, value=formula)
    ws2.cell(row=4, column=col).font = Font(name="Georgia", bold=True, size=24, color=MIDNIGHT)
    ws2.cell(row=4, column=col).alignment = Alignment(horizontal="center")
    if label == "Avg Adoption":
        ws2.cell(row=4, column=col).number_format = "0%"

    # Card border
    for row in [3, 4]:
        ws2.cell(row=row, column=col).border = thin_border

ws2.row_dimensions[4].height = 40

# Category breakdown table
ws2.cell(row=6, column=1, value="Category Breakdown")
ws2.cell(row=6, column=1).font = Font(name="Arial", bold=True, size=12, color=MIDNIGHT)

cat_headers = ["Category", "Count", "Avg Adoption"]
for i, h in enumerate(cat_headers, 1):
    ws2.cell(row=7, column=i, value=h)
style_header_row(ws2, 7, 3)

categories = ["Document Gen", "Design", "QA/Testing", "Debugging", "Deployment",
              "Code Review", "Integration", "Strategy", "Team", "Testing"]

for r, cat in enumerate(categories, 8):
    ws2.cell(row=r, column=1, value=cat)
    ws2.cell(row=r, column=2, value=f"=COUNTIF('Skills Inventory'!B:B,A{r})")
    ws2.cell(row=r, column=3, value=f"=AVERAGEIF('Skills Inventory'!B:B,A{r},'Skills Inventory'!F:F)")
    ws2.cell(row=r, column=3).number_format = "0%"
    style_data_row(ws2, r, 3, alternate=(r % 2 == 0))

# ============================================================
# SHEET 3 — Adoption Timeline
# ============================================================
ws3 = wb.create_sheet("Adoption Timeline")

headers3 = ["Tier", "Timeline", "Skills Count", "Target Adoption", "Priority"]
col_widths3 = [14, 20, 16, 18, 14]

for i, (header, width) in enumerate(zip(headers3, col_widths3), 1):
    ws3.cell(row=1, column=i, value=header)
    ws3.column_dimensions[get_column_letter(i)].width = width

style_header_row(ws3, 1, len(headers3))

timeline_data = [
    ["Day 1", "Immediate", f"=COUNTIF('Skills Inventory'!C:C,\"Day 1\")", 0.90, "Critical"],
    ["Week 1", "First week", f"=COUNTIF('Skills Inventory'!C:C,\"Week 1\")", 0.75, "High"],
    ["Week 2-3", "Weeks 2-3", f"=COUNTIF('Skills Inventory'!C:C,\"Week 2-3\")", 0.60, "Medium"],
    ["Week 4+", "Month 2+", f"=COUNTIF('Skills Inventory'!C:C,\"Week 4+\")", 0.40, "Low"],
]

for r, row_data in enumerate(timeline_data, 2):
    for c, value in enumerate(row_data, 1):
        ws3.cell(row=r, column=c, value=value)
    ws3.cell(row=r, column=4).number_format = "0%"
    style_data_row(ws3, r, len(headers3), alternate=(r % 2 == 0))

# Conditional formatting on Priority
ws3.conditional_formatting.add(
    "E2:E100",
    CellIsRule(operator="equal", formula=['"Critical"'],
              fill=PatternFill("solid", fgColor="FEE2E2"),
              font=Font(color=RED, bold=True))
)
ws3.conditional_formatting.add(
    "E2:E100",
    CellIsRule(operator="equal", formula=['"High"'],
              fill=PatternFill("solid", fgColor="FEF3C7"),
              font=Font(color=YELLOW, bold=True))
)

ws3.freeze_panes = "A2"

# Total row
total_row = len(timeline_data) + 2
ws3.cell(row=total_row, column=1, value="TOTAL")
ws3.cell(row=total_row, column=1).font = Font(name="Arial", bold=True, size=10)
ws3.cell(row=total_row, column=3, value=f"=SUM(C2:C{total_row - 1})")
ws3.cell(row=total_row, column=3).font = Font(name="Arial", bold=True, size=10)
for c in range(1, len(headers3) + 1):
    ws3.cell(row=total_row, column=c).border = Border(
        top=Side(style="double", color=MIDNIGHT),
        bottom=Side(style="double", color=MIDNIGHT),
    )

# --- Save ---
output = "ai_skills_adoption_tracker.xlsx"
wb.save(output)
size_kb = os.path.getsize(output) / 1024
print(f"Created: {output} ({size_kb:.1f} KB)")
print(f"Sheets: {', '.join(wb.sheetnames)}")
